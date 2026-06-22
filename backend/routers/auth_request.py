import asyncio
import base64
import os
import re
import tempfile

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session as DBSession

from audit_log import write_log
from auth_utils import get_current_user
from config import AUTH_LEDGER_PATH
from db import get_db
from integrations.dingtalk import notify_task_failure, notify_task_success
from models import User
from perf_trace import PerfTrace
from routers.chat import load_history, save_history
from task_runner import create_background_task, submit_background_task
from upload_validation import UploadValidationError, validate_excel_upload, validate_legal_upload, validate_pdf_upload
from utils.ledger_importer import (
    create_pending_import,
    load_pending_import,
    make_preview,
    normalize_key,
    parse_auth_rows,
)

router = APIRouter(prefix="/api/auth-request", tags=["auth-request"])


class AuthPreviewRequest(BaseModel):
    extracted: dict
    user_inputs: dict


class AuthGenerateRequest(BaseModel):
    extracted: dict
    user_inputs: dict
    content: str
    session_id: str = ""


class AuthLedgerRecordRequest(BaseModel):
    extracted: dict
    user_inputs: dict
    title: str
    session_id: str = ""


class AuthLedgerImportConfirmRequest(BaseModel):
    import_token: str
    session_id: str = ""


def _auth_existing_keys() -> set[str]:
    import openpyxl
    from utils.auth_request_drafter import AUTH_LEDGER_HEADERS

    if not os.path.exists(AUTH_LEDGER_PATH):
        return set()
    try:
        wb = openpyxl.load_workbook(AUTH_LEDGER_PATH, data_only=True)
        ws = wb.active
        auth_no_col = AUTH_LEDGER_HEADERS.index("编号") + 1
        keys = set()
        for row_idx in range(2, ws.max_row + 1):
            auth_no = normalize_key(ws.cell(row_idx, auth_no_col).value)
            if auth_no:
                keys.add(f"编号:{auth_no}")
        return keys
    except Exception:
        return set()


def _auth_import_key(row: dict) -> str:
    auth_no = normalize_key(row.get("编号"))
    if auth_no:
        return f"编号:{auth_no}"
    return ""


def _ocr_pdf_if_needed(pdf_bytes: bytes, text: str, vision_model: str) -> str:
    if text.strip():
        return text
    from ledger_helpers import ocr_pdf_with_vision

    return ocr_pdf_with_vision(pdf_bytes, model=vision_model)


def _safe_filename(name: str, suffix: str = ".docx") -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", name).strip()
    cleaned = re.sub(r"\s+", "", cleaned)
    return f"{cleaned[:120] or '授权请示'}{suffix}"


def _run_auth_extract(
    attachment1_bytes: bytes,
    attachment2_bytes: bytes,
    attachment2_filename: str,
    vision_model: str,
    user_id: int,
    *,
    on_progress=None,
) -> dict:
    from utils.auth_request_drafter import (
        extract_attachment1_info,
        extract_attachment2_info,
        extract_attachment_text,
    )

    trace = PerfTrace("auth_request.extract", user_id)

    def progress(value: int, message: str) -> None:
        if on_progress:
            on_progress(value, message)

    try:
        progress(20, "正在解析依据文件")
        with trace.step("attachment1_extract"):
            a1 = extract_attachment1_info(attachment1_bytes)
            if not a1.get("raw_text"):
                a1 = extract_attachment1_info(
                    attachment1_bytes,
                    ocr_text=_ocr_pdf_if_needed(attachment1_bytes, "", vision_model),
                )

        progress(55, "正在解析授权委托书")
        with trace.step("attachment2_extract"):
            ext = os.path.splitext(attachment2_filename)[1].lower()
            if ext == ".pdf":
                text = extract_attachment_text(attachment2_bytes, attachment2_filename)
                ocr_text = _ocr_pdf_if_needed(attachment2_bytes, text, vision_model)
                a2 = extract_attachment2_info(attachment2_bytes, attachment2_filename, ocr_text=ocr_text)
            else:
                a2 = extract_attachment2_info(attachment2_bytes, attachment2_filename)

        progress(90, "字段提取完成")
        return {"attachment1": a1, "attachment2": a2}
    finally:
        trace.finish()


@router.post("/extract-task")
async def start_auth_extract_task(
    attachment1_file: UploadFile = File(...),
    attachment2_file: UploadFile = File(...),
    session_id: str = Form(""),
    vision_model: str = Form(""),
    db: DBSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    attachment1_bytes = await attachment1_file.read()
    attachment2_bytes = await attachment2_file.read()
    try:
        validate_pdf_upload(attachment1_file.filename or "", attachment1_file.content_type, attachment1_bytes)
        attachment2_name = validate_legal_upload(attachment2_file.filename or "", attachment2_file.content_type, attachment2_bytes)
    except UploadValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))

    task = create_background_task(
        db,
        task_type="auth_request_extract",
        created_by=user.id,
        message="已接收附件，等待提取",
    )

    def _worker(ctx, task_db: DBSession) -> dict:
        task_user = task_db.query(User).filter(User.id == user.id).first()
        if not task_user:
            raise RuntimeError("发起用户不存在")
        try:
            extracted = _run_auth_extract(
                attachment1_bytes,
                attachment2_bytes,
                attachment2_name,
                vision_model,
                task_user.id,
                on_progress=lambda progress, message: ctx.update(progress=progress, message=message),
            )
            write_log(task_db, task_user, "auth_request_extract", "提取授权请示附件字段", None)
            notify_task_success(
                task="授权请示字段提取",
                summary=(extracted.get("attachment1") or {}).get("project_name") or "授权请示",
                user=task_user,
                session_id=session_id,
                stage="字段提取",
            )
            return {"extracted": extracted}
        except Exception as exc:
            notify_task_failure(
                task="授权请示字段提取",
                summary=str(exc)[:160],
                user=task_user,
                session_id=session_id,
                stage="字段提取",
            )
            raise

    submit_background_task(task.task_id, _worker)
    return {"ok": True, "task_id": task.task_id}


@router.post("/preview")
def preview_auth_request(body: AuthPreviewRequest):
    from utils.auth_request_drafter import build_auth_request_text, validate_user_inputs

    try:
        validate_user_inputs(body.user_inputs)
        content = build_auth_request_text(body.extracted, body.user_inputs)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"content": content}


@router.post("/generate")
async def generate_auth_request_docx(
    body: AuthGenerateRequest,
    request: Request,
    db: DBSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from utils.auth_request_drafter import save_auth_request_docx, validate_user_inputs

    try:
        validate_user_inputs(body.user_inputs)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    raw_project_name = str((body.extracted.get("attachment1") or {}).get("project_name") or "").strip()
    project_name = raw_project_name if raw_project_name and len(raw_project_name) <= 80 else "授权请示"
    title = f"关于办理{project_name}授权委托书的请示"

    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False, prefix="授权请示_") as tmp:
        docx_path = tmp.name
    try:
        await asyncio.to_thread(save_auth_request_docx, body.content, docx_path)
        with open(docx_path, "rb") as f:
            docx_b64 = base64.b64encode(f.read()).decode()
    finally:
        try:
            os.unlink(docx_path)
        except OSError:
            pass

    write_log(db, user, "auth_request_generate", f"生成授权请示：{project_name}", request)
    reply = f"✅ 授权请示已生成：{title}"
    if body.session_id:
        history = await asyncio.to_thread(load_history, user.id, body.session_id)
        history.append({"role": "assistant", "content": reply})
        await asyncio.to_thread(save_history, history, user.id, body.session_id)

    return {
        "content": body.content,
        "docx_base64": docx_b64,
        "filename": _safe_filename(title),
        "title": title,
        "ledger_updated": False,
        "ledger_base64": None,
        "ledger_filename": None,
    }


@router.post("/record-ledger")
def record_auth_ledger(
    body: AuthLedgerRecordRequest,
    request: Request,
    db: DBSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from utils.auth_request_drafter import record_to_ledger, validate_user_inputs

    try:
        validate_user_inputs(body.user_inputs)
        ledger_updated = record_to_ledger(body.extracted, body.user_inputs, body.title, AUTH_LEDGER_PATH)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    ledger_b64 = None
    ledger_filename = None
    if ledger_updated:
        with open(AUTH_LEDGER_PATH, "rb") as f:
            ledger_b64 = base64.b64encode(f.read()).decode()
        ledger_filename = os.path.basename(AUTH_LEDGER_PATH)

    write_log(db, user, "auth_request_ledger_write", f"记录授权台账：{body.title}", request)
    reply = f"✅ 授权委托台账已记录：{body.title}"
    if body.session_id:
        history = load_history(user.id, body.session_id)
        history.append({"role": "assistant", "content": reply})
        save_history(history, user.id, body.session_id)
    return {
        "ledger_updated": ledger_updated,
        "ledger_base64": ledger_b64,
        "ledger_filename": ledger_filename,
        "reply": reply,
    }


@router.post("/import-ledger-preview")
async def preview_auth_ledger_import(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
):
    from utils.auth_request_drafter import AUTH_LEDGER_HEADERS

    file_bytes = await file.read()
    try:
        validate_excel_upload(file.filename or "", file.content_type, file_bytes)
        records, invalid_rows = parse_auth_rows(file_bytes, AUTH_LEDGER_HEADERS)
    except (UploadValidationError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    token = create_pending_import(user.id, "auth_ledger", {"records": records, "invalid_rows": invalid_rows})
    return make_preview(
        records=records,
        invalid_rows=invalid_rows,
        existing_keys=_auth_existing_keys(),
        key_fn=_auth_import_key,
        import_token=token,
    )


@router.post("/import-ledger-confirm")
def confirm_auth_ledger_import(
    body: AuthLedgerImportConfirmRequest,
    request: Request,
    db: DBSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from utils.auth_request_drafter import import_auth_ledger_rows

    try:
        pending = load_pending_import(user.id, "auth_ledger", body.import_token)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    result = import_auth_ledger_rows(pending.get("records") or [], AUTH_LEDGER_PATH)
    reply = f"✅ 历史授权委托台账导入完成：新增 {result['inserts']} 条，更新 {result['updates']} 条。"
    if body.session_id:
        history = load_history(user.id, body.session_id)
        history.append({"role": "assistant", "content": reply})
        save_history(history, user.id, body.session_id)
    write_log(db, user, "auth_request_ledger_import", f"导入历史授权台账：新增 {result['inserts']}，更新 {result['updates']}", request)
    return {"ok": True, **result, "reply": reply}


# Backward-compatible endpoint name for older frontend bundles. It now starts
# the new two-attachment extraction task.
@router.post("/process-task")
async def start_auth_request_process_task(
    attachment1_file: UploadFile | None = File(None),
    attachment2_file: UploadFile | None = File(None),
    pdf_file: UploadFile | None = File(None),
    session_id: str = Form(""),
    vision_model: str = Form(""),
    db: DBSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if not attachment1_file or not attachment2_file:
        if pdf_file is not None:
            raise HTTPException(status_code=400, detail="请同时上传依据文件和授权委托书")
        raise HTTPException(status_code=400, detail="缺少依据文件或授权委托书")
    return await start_auth_extract_task(attachment1_file, attachment2_file, session_id, vision_model, db, user)
