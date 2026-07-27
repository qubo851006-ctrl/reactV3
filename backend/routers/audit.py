# -*- coding: utf-8 -*-
import asyncio
import io
import json
import logging
import re
import tempfile

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session as DBSession

from auth_utils import get_current_user
from audit_log import write_log
from db import get_db
from integrations.dingtalk import notify_task_failure, notify_task_success
from models import User
from llm_client import get_llm_client
from task_runner import create_background_task, submit_background_task
from upload_validation import UploadValidationError, validate_excel_upload
from perf_trace import PerfTrace

router = APIRouter(prefix="/api/audit")
AUDIT_CLASSIFY_MODEL = "qwen3.6"
AUDIT_REVIEW_MODEL = "deepseek-v4-flash"


# ── 分类体系常量 ───────────────────────────────────────────────────

CATEGORY_TAXONOMY: dict[str, list[str]] = {
    "公司治理": ["三重一大决策", "董事会管理", "股东会管理", "会议管理", "其他"],
    "合同及法律合规管理": ["合同审核及签署", "合同执行", "诉讼管理", "授权管理", "知识产权管理", "其他"],
    "采购管理": ["采购方式", "采购评审", "供应商管理", "采购文档管理", "其他"],
    "营销管理": ["代理人管理", "销售管理", "大客户管理", "团队管理", "常旅客管理", "知音商城管理", "品牌管理", "其他"],
    "人力资源管理": ["人员招聘", "绩效考核", "薪酬福利", "考勤管理", "培训管理", "岗位与人员配置", "离职管理", "领导人员履职待遇", "其他"],
    "财务管理": ["预算管理", "银行账户管理", "成本费用管理", "资金管理", "往来账款管理", "会计核算管理", "保险及索赔管理", "担保管理", "税务管理", "优惠政策使用管理", "其他"],
    "资产管理": ["固定资产管理", "存货管理", "低值易耗品管理", "无形资产管理", "资产权证管理", "其他"],
    "信息系统管理": ["系统功能开发管理", "系统账号管理", "系统安全管理", "系统应用管理", "其他"],
    "工程项目管理": ["工程项目工期管理", "工程项目招标管理", "工程项目洽商变更", "施工过程管理", "工程项目验收", "工程项目竣工结决算", "其他"],
    "安全管理": ["安全事件", "安全与质量考核", "空防、消防、地面安全管理", "应急管理", "其他"],
    "内部控制管理": ["评价管理", "内部控制手册建设", "风险识别与管理", "规章制度审核", "其他"],
    "行政管理（其他）": ["中央八项规定精神", "档案管理", "证照管理", "礼品管理", "印章管理", "免折票管理", "审计整改", "对外捐赠", "企业文化"],
    "其他": ["其他"],
}


# ── 数据模型 ──────────────────────────────────────────────────────


class Disagreement(BaseModel):
    category_l1: str
    category_l2: str
    domain: str


class AuditRow(BaseModel):
    seq: int
    issue: str
    description: str
    category_l1: str = ""
    category_l2: str = ""
    domain: str = ""
    disagreement: Disagreement | None = None  # B 的修正建议，None 表示 A/B 一致


class DownloadRequest(BaseModel):
    rows: list[AuditRow]
    original_filename: str = "审计问题分析结果"


class AuditClassifyItem(BaseModel):
    seq: int = Field(alias="序号")
    category_l1: str = Field(alias="问题类别一级")
    category_l2: str = Field(alias="问题类别二级")
    domain: str = Field(alias="业务领域")


class AuditClassifyOutput(BaseModel):
    items: list[AuditClassifyItem]


_JSON_ONLY_SYSTEM_MESSAGE = (
    "You are a JSON-only API. Do not output reasoning, thinking process, "
    "markdown, explanations, or code fences. Return only one valid JSON object."
)


# ── 工具函数 ──────────────────────────────────────────────────────


def _find_header_row(ws) -> int | None:
    """找到表头行：某单元格值精确等于「发现问题」的行（1-based）。
    使用精确匹配避免把含「发现问题」子串的标题行（如「审计发现问题汇总表」）误判为表头。
    """
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row_idx, col_idx).value
            if cell_val and str(cell_val).strip() == "发现问题":
                return row_idx
    return None


def _col_index(ws, header_row: int, keyword: str) -> int | None:
    """在标题行中找含 keyword 的列号（1-based）。"""
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col_idx).value
        if val and keyword in str(val):
            return col_idx
    return None


def _extract_rows(wb: openpyxl.Workbook) -> list[dict]:
    """提取第一个 sheet 中的审计发现行数据。"""
    ws = wb.active
    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError("未找到包含「发现问题」的标题行，请检查 Excel 格式。")

    seq_col = _col_index(ws, header_row, "序号") or 1
    issue_col = _col_index(ws, header_row, "发现问题")
    desc_col = _col_index(ws, header_row, "问题描述")

    if issue_col is None:
        raise ValueError("未找到「发现问题」列。")

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        seq_val = ws.cell(row_idx, seq_col).value
        issue_val = ws.cell(row_idx, issue_col).value
        desc_val = ws.cell(row_idx, desc_col).value if desc_col else ""

        if not issue_val:
            continue

        try:
            seq = int(seq_val) if seq_val is not None else row_idx - header_row
        except (ValueError, TypeError):
            seq = row_idx - header_row

        rows.append({
            "seq": seq,
            "issue": str(issue_val).strip(),
            "description": str(desc_val).strip() if desc_val else "",
        })

    return rows


def _build_prompt(rows: list[dict], domains: list[str]) -> str:
    rows_json = json.dumps(
        [{"序号": r["seq"], "发现问题": r["issue"], "问题描述": r["description"][:200]}
         for r in rows],
        ensure_ascii=False,
        indent=2,
    )

    # 构建一级→二级对应关系说明
    taxonomy_lines = "\n".join(
        f"- {l1}：{'/ '.join(l2_list)}"
        for l1, l2_list in CATEGORY_TAXONOMY.items()
    )

    domain_lines = "\n".join(f"- {d}" for d in domains)

    return f"""你是企业内部审计专家。请对以下审计发现问题进行分类，共两个独立维度，分别判断。

【维度一：问题类别（两级）】
请先选择最匹配的一级类别，再在该一级类别下选择最匹配的二级子类。

一级类别及其对应二级子类：
{taxonomy_lines}

【维度二：业务领域】
描述问题所属的业务板块，从以下选项中选一个最匹配的：
{domain_lines}

【发现问题列表（JSON）】：
{rows_json}

请严格按如下格式返回，只输出一个合法 JSON 对象，不含任何其他文字、解释、Markdown 或思考过程：
{{"items": [{{"序号": 1, "问题类别一级": "...", "问题类别二级": "...", "业务领域": "..."}}]}}"""


def _iter_json_payloads(text: str):
    """Yield valid JSON object/array candidates from possibly noisy LLM output."""
    cleaned = (text or "").strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)

    decoder = json.JSONDecoder()
    for match in re.finditer(r"[\{\[]", cleaned):
        try:
            value, _ = decoder.raw_decode(cleaned[match.start():])
        except json.JSONDecodeError:
            continue
        if isinstance(value, (dict, list)):
            yield value


def _extract_json_payload(text: str) -> object:
    """Extract the first valid JSON object/array from possibly noisy LLM output."""
    for payload in _iter_json_payloads(text):
        return payload
    raise ValueError("LLM 返回格式异常，无法解析 JSON：未找到 JSON 对象或数组")


def _coerce_classify_output(payload: object) -> AuditClassifyOutput:
    if isinstance(payload, list):
        payload = {"items": payload}
    if not isinstance(payload, dict):
        raise ValueError("LLM 返回格式异常，JSON 顶层必须是对象。")
    try:
        return AuditClassifyOutput.model_validate(payload)
    except Exception as exc:
        raise ValueError(f"LLM 返回格式异常，分类结果字段不完整或类型错误：{str(exc)[:160]}") from exc


def _extract_classify_output(text: str) -> AuditClassifyOutput:
    errors: list[str] = []
    for payload in _iter_json_payloads(text):
        try:
            return _coerce_classify_output(payload)
        except ValueError as exc:
            errors.append(str(exc))
            continue
    detail = errors[-1] if errors else "未找到 JSON 对象或数组"
    raise ValueError(f"LLM 返回格式异常，无法解析审计分类 JSON：{detail}")


def _normalise_classify_item(
    item: AuditClassifyItem,
    *,
    valid_domains: set[str] | None,
) -> dict:
    category_l1 = (item.category_l1 or "").strip()
    category_l2 = (item.category_l2 or "").strip()
    domain = (item.domain or "").strip()

    if category_l1 not in CATEGORY_TAXONOMY:
        category_l1 = ""
        category_l2 = ""
    elif category_l2 not in CATEGORY_TAXONOMY[category_l1]:
        category_l2 = ""

    if valid_domains is not None and domain not in valid_domains:
        domain = ""

    return {
        "category_l1": category_l1,
        "category_l2": category_l2,
        "domain": domain,
    }


def _parse_llm_output(
    text: str,
    rows: list[dict],
    domains: list[str] | None = None,
) -> list[dict]:
    """Parse structured audit classifications and fill missing rows."""
    output = _extract_classify_output(text)
    valid_domains = set(domains) if domains else None
    result_map = {
        item.seq: _normalise_classify_item(item, valid_domains=valid_domains)
        for item in output.items
    }
    if rows and not any(
        item["category_l1"] and item["category_l2"] and item["domain"]
        for item in result_map.values()
    ):
        raise ValueError("LLM 返回格式异常，未提取到可用完整审计分类结果。")

    result = []
    for r in rows:
        classified = result_map.get(r["seq"], {})
        result.append({
            **r,
            "category_l1": classified.get("category_l1", ""),
            "category_l2": classified.get("category_l2", ""),
            "domain": classified.get("domain", ""),
        })
    return result


def _build_classify_messages(prompt: str) -> list[dict]:
    return [
        {"role": "system", "content": _JSON_ONLY_SYSTEM_MESSAGE},
        {"role": "user", "content": prompt},
    ]


def _build_repair_prompt(raw_output: str, rows: list[dict], domains: list[str], error: str) -> str:
    seqs = [r["seq"] for r in rows]
    domain_lines = "\n".join(f"- {d}" for d in domains)
    return f"""上一次审计分类模型输出不是合法 JSON，解析错误为：{error}

请只修复格式，不要重新解释过程。根据下面的模型原始输出，提取已有分类结果并转换为一个合法 JSON 对象。

必须返回这个 schema：
{{"items": [{{"序号": 1, "问题类别一级": "...", "问题类别二级": "...", "业务领域": "..."}}]}}

要求：
- 只返回 JSON 对象，不要 Markdown、解释、思考过程或代码块
- "序号" 只能来自这些值：{seqs}
- "业务领域" 只能来自：
{domain_lines}
- 如果原始输出里没有可用分类，返回 {{"items": []}}

模型原始输出：
{raw_output[:20000]}"""


def _call_classify_llm(client, prompt: str, rows: list[dict], domains: list[str]) -> list[dict]:
    from llm_audit import traced_complete

    resp = traced_complete(
        client,
        scene="audit_classify",
        prompt_template_id="audit.classify.v2",
        model=AUDIT_CLASSIFY_MODEL,
        messages=_build_classify_messages(prompt),
        temperature=0,
    )
    raw_output = resp.choices[0].message.content or ""
    try:
        return _parse_llm_output(raw_output, rows, domains)
    except ValueError as first_error:
        first_error_text = str(first_error)
        logging.warning(
            "audit_classify JSON parse failed; retrying format repair: %s",
            first_error_text[:200],
        )

    repair_resp = traced_complete(
        client,
        scene="audit_classify_repair",
        prompt_template_id="audit.classify.repair.v1",
        model=AUDIT_CLASSIFY_MODEL,
        messages=[
            {"role": "system", "content": _JSON_ONLY_SYSTEM_MESSAGE},
            {"role": "user", "content": _build_repair_prompt(raw_output, rows, domains, first_error_text)},
        ],
        temperature=0,
    )
    repair_output = repair_resp.choices[0].message.content or ""
    try:
        return _parse_llm_output(repair_output, rows, domains)
    except ValueError as repair_error:
        logging.warning(
            "audit_classify repair JSON parse failed: %s",
            str(repair_error)[:200],
        )
        raise ValueError(
            "模型返回格式异常，已尝试修复但仍无法解析。请减少单次审计问题数量后重试，或联系管理员查看 LLM 审计记录。"
        ) from repair_error


def _build_review_prompt(rows_a: list[dict], domains: list[str]) -> str:
    """构建模型B的审查提示词：逐条质检模型A的分类结果。"""
    taxonomy_lines = "\n".join(
        f"- {l1}：{'/ '.join(l2_list)}"
        for l1, l2_list in CATEGORY_TAXONOMY.items()
    )
    domain_lines = "\n".join(f"- {d}" for d in domains)
    rows_json = json.dumps(
        [{"序号": r["seq"], "发现问题": r["issue"],
          "已分类一级": r["category_l1"], "已分类二级": r["category_l2"],
          "已分类领域": r["domain"]}
         for r in rows_a],
        ensure_ascii=False,
        indent=2,
    )
    return f"""你是企业内部审计问题分类质检专家。请逐条审查以下AI分类结果是否准确。

【分类体系（一级→二级）】
{taxonomy_lines}

【可用业务领域】
{domain_lines}

【待审查的分类结果】
{rows_json}

请严格按如下格式返回，只输出JSON数组，不含任何其他文字：
[{{"序号": 1, "需要修正": false, "问题类别一级": "...", "问题类别二级": "...", "业务领域": "..."}}]

规则：
- "需要修正"为false时，后三个字段填写与原分类相同的值
- "需要修正"为true时，填写你认为更准确的分类（必须来自给定分类体系）
- 所有字段不可为空"""


def _call_review_llm(rows_a: list[dict], domains: list[str]) -> list[dict]:
    """调用模型B对A的分类结果进行逐条质检，返回需修正的行列表。"""
    import logging
    from llm_audit import traced_complete
    client = get_llm_client()
    prompt = _build_review_prompt(rows_a, domains)
    resp = traced_complete(
        client,
        scene="audit_cross_review",
        prompt_template_id="audit.cross_review.v1",
        model=AUDIT_REVIEW_MODEL,
        messages=[
            {"role": "system", "content": _JSON_ONLY_SYSTEM_MESSAGE},
            {"role": "user", "content": prompt},
        ],
        temperature=0,
    )
    text = resp.choices[0].message.content or ""
    try:
        payload = _extract_json_payload(text)
        if isinstance(payload, dict):
            parsed = payload.get("items", [])
        else:
            parsed = payload
        if not isinstance(parsed, list):
            logging.warning("模型B返回JSON不是数组，原始文本: %s", text[:200])
            return []
    except ValueError:
        logging.warning("模型B返回JSON解析失败，原始文本: %s", text[:200])
        return []
    return [item for item in parsed if item.get("需要修正")]


def _merge_ab_results(rows_a: list[dict], corrections: list[dict]) -> list[dict]:
    """将 B 的修正意见合并进 A 的结果，加入 disagreement 字段。"""
    correction_map = {item["序号"]: item for item in corrections}
    result = []
    for row in rows_a:
        corr = correction_map.get(row["seq"])
        merged = dict(row)
        if corr:
            merged["disagreement"] = {
                "category_l1": corr.get("问题类别一级", ""),
                "category_l2": corr.get("问题类别二级", ""),
                "domain": corr.get("业务领域", ""),
            }
        else:
            merged["disagreement"] = None
        result.append(merged)
    return result


def _run_audit_analysis(
    content: bytes,
    filename: str,
    content_type: str | None,
    doms: list[str],
    user_id: int,
    *,
    on_progress=None,
) -> tuple[list[dict], list[str]]:
    trace = PerfTrace("audit.analyze", user_id)

    def progress(value: int, message: str) -> None:
        if on_progress:
            on_progress(value, message)

    try:
        progress(10, "正在读取审计 Excel")
        validate_excel_upload(filename or "", content_type, content)
        with trace.step("load_workbook"):
            wb = openpyxl.load_workbook(io.BytesIO(content))
        with trace.step("extract_rows"):
            rows = _extract_rows(wb)
        if not rows:
            raise ValueError("Excel 中未找到有效数据行。")

        prompt = _build_prompt(rows, doms)

        from llm_audit.context import collect_traces
        client = get_llm_client()

        with collect_traces() as bucket:
            progress(35, "模型 A 正在初步分类")
            with trace.step("model_a_classify"):
                rows_a = _call_classify_llm(client, prompt, rows, doms)

            progress(70, "模型 B 正在交叉校验")
            with trace.step("model_b_review"):
                try:
                    corrections = _call_review_llm(rows_a, doms)
                except Exception:
                    corrections = []

            progress(90, "正在合并分类结果")
            with trace.step("merge_results"):
                merged = _merge_ab_results(rows_a, corrections)
            return merged, list(bucket.ids)
    finally:
        trace.finish()


# ── 路由 ──────────────────────────────────────────────────────────


@router.post("/analyze")
async def analyze_audit(
    file: UploadFile = File(...),
    domains: str = Form('["物业租赁","酒店公寓","工程领域","资产处置","历史遗留问题"]'),
    request: Request = None,
    db: DBSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    try:
        doms = json.loads(domains)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"业务领域参数格式错误：{e}")

    trace = PerfTrace("audit.analyze", user.id)
    try:
        content = await file.read()
        validate_excel_upload(file.filename or "", file.content_type, content)
        with trace.step("load_workbook"):
            wb = openpyxl.load_workbook(io.BytesIO(content))
        with trace.step("extract_rows"):
            rows = _extract_rows(wb)
    except UploadValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败：{e}")

    if not rows:
        raise HTTPException(status_code=400, detail="Excel 中未找到有效数据行。")

    prompt = _build_prompt(rows, doms)

    def _run_full_analysis() -> tuple[list, list[str]]:
        from llm_audit.context import collect_traces
        client = get_llm_client()

        with collect_traces() as bucket:
            # Step 1: 模型 A 固定用 Qwen，避免全局默认模型影响审计分类。
            with trace.step("model_a_classify"):
                rows_a = _call_classify_llm(client, prompt, rows, doms)

            # Step 2: 模型 B 固定用 DeepSeek 逐条审查 A 的结果。
            with trace.step("model_b_review"):
                try:
                    corrections = _call_review_llm(rows_a, doms)
                except Exception:
                    corrections = []  # B 失败静默降级，只返回 A 的结果

            # Step 3: 合并差异信息
            with trace.step("merge_results"):
                merged = _merge_ab_results(rows_a, corrections)
            return merged, list(bucket.ids)

    try:
        classified_rows, llm_trace_ids = await asyncio.to_thread(_run_full_analysis)
    except ValueError as e:
        notify_task_failure(
            task="审计问题分析",
            summary=str(e)[:160],
            user=user,
            stage="数据校验",
        )
        raise HTTPException(status_code=502, detail=str(e))
    except Exception as e:
        notify_task_failure(
            task="审计问题分析",
            summary=str(e)[:160],
            user=user,
            stage="AI 分析",
        )
        raise HTTPException(status_code=502, detail=f"LLM 调用失败：{e}")
    finally:
        trace.finish()

    disagreement_count = sum(1 for r in classified_rows if r.get("disagreement"))
    write_log(
        db, user, "audit_analyze",
        f"审计分析 {len(classified_rows)} 条，其中 {disagreement_count} 条存在分类分歧",
        request,
    )
    notify_task_success(
        task="审计问题分析",
        summary=f"共 {len(classified_rows)} 条，{disagreement_count} 条存在分类分歧",
        user=user,
        stage="AI 分析",
    )
    return {"rows": classified_rows, "total": len(classified_rows), "llm_trace_ids": llm_trace_ids}


@router.post("/analyze-task")
async def start_audit_analyze_task(
    file: UploadFile = File(...),
    domains: str = Form('["物业租赁","酒店公寓","工程领域","资产处置","历史遗留问题"]'),
    db: DBSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    try:
        doms = json.loads(domains)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"业务领域参数格式错误：{e}")

    content = await file.read()
    try:
        validate_excel_upload(file.filename or "", file.content_type, content)
    except UploadValidationError as e:
        raise HTTPException(status_code=400, detail=str(e))

    filename = file.filename or ""
    content_type = file.content_type
    task = create_background_task(
        db,
        task_type="audit_analyze",
        created_by=user.id,
        message="已接收文件，等待分析",
    )

    def _worker(ctx, task_db: DBSession) -> dict:
        task_user = task_db.query(User).filter(User.id == user.id).first()
        if not task_user:
            raise RuntimeError("发起用户不存在")
        try:
            classified_rows, llm_trace_ids = _run_audit_analysis(
                content,
                filename,
                content_type,
                doms,
                task_user.id,
                on_progress=lambda progress, message: ctx.update(progress=progress, message=message),
            )
            disagreement_count = sum(1 for r in classified_rows if r.get("disagreement"))
            write_log(
                task_db,
                task_user,
                "audit_analyze",
                f"审计分析 {len(classified_rows)} 条，其中 {disagreement_count} 条存在分类分歧",
                None,
            )
            notify_task_success(
                task="审计问题分析",
                summary=f"共 {len(classified_rows)} 条，{disagreement_count} 条存在分类分歧",
                user=task_user,
                stage="AI 分析",
            )
            return {"rows": classified_rows, "total": len(classified_rows), "llm_trace_ids": llm_trace_ids}
        except Exception as exc:
            notify_task_failure(
                task="审计问题分析",
                summary=str(exc)[:160],
                user=task_user,
                stage="AI 分析",
            )
            raise

    submit_background_task(task.task_id, _worker)
    return {"ok": True, "task_id": task.task_id}


@router.post("/download")
async def download_audit_excel(req: DownloadRequest):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审计问题分析"

    # 样式定义
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    headers = ["序号", "发现问题", "问题描述", "问题类别一级", "问题类别二级", "业务领域"]
    col_widths = [8, 30, 50, 18, 20, 14]

    # 写标题行
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20

    # 分类列填充色
    cat_l1_fill = PatternFill("solid", fgColor="EBF5FB")
    cat_l2_fill = PatternFill("solid", fgColor="D6EAF8")
    dom_fill = PatternFill("solid", fgColor="E9F7EF")

    for row_idx, row in enumerate(req.rows, start=2):
        ws.cell(row_idx, 1, row.seq).alignment = center_align
        ws.cell(row_idx, 2, row.issue).alignment = wrap_align
        ws.cell(row_idx, 3, row.description).alignment = wrap_align
        l1_cell = ws.cell(row_idx, 4, row.category_l1)
        l1_cell.alignment = center_align
        l1_cell.fill = cat_l1_fill
        l2_cell = ws.cell(row_idx, 5, row.category_l2)
        l2_cell.alignment = center_align
        l2_cell.fill = cat_l2_fill
        dom_cell = ws.cell(row_idx, 6, row.domain)
        dom_cell.alignment = center_align
        dom_cell.fill = dom_fill

    # 冻结标题行
    ws.freeze_panes = "A2"

    # 写入临时文件并返回
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)

    safe_name = req.original_filename.replace("/", "_").replace("\\", "_")
    return FileResponse(
        path=tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{safe_name}_分类结果.xlsx",
    )
