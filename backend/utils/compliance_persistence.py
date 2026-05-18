"""Compliance ledger persistence: JSON read/write + Excel workbook builder.

Extracted from compliance_ledger.py so the parent module can focus on
PDF extraction + LLM review logic. Public surface (load_records,
save_records, upsert_record, append_record, create_compliance_workbook)
is re-exported from utils.compliance_ledger so existing imports keep
working.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import COMPLIANCE_LEDGER_EXCEL_PATH, COMPLIANCE_LEDGER_JSON_PATH
from file_store import atomic_write_text, file_lock


WORKSHEET_NAME = "合规管理牵头部门合规审查台账"
HEADERS = [
    "序号", "重大事项", "程序", "审查时间", "审查单位",
    "合规审查意见", "具体意见描述", "落实情况", "承办单位", "背景材料",
]


def load_records(path: str | Path = COMPLIANCE_LEDGER_JSON_PATH) -> list[dict[str, Any]]:
    p = Path(path)
    if not p.exists():
        return []
    try:
        with file_lock(p):
            data = json.loads(p.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    return data if isinstance(data, list) else []


def save_records(records: list[dict[str, Any]], path: str | Path = COMPLIANCE_LEDGER_JSON_PATH) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with file_lock(p):
        atomic_write_text(p, json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_record_key(value: Any) -> str:
    return "".join(str(value or "").split()).lower()


def upsert_record(
    records: list[dict[str, Any]],
    record: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any], bool]:
    """Insert or update by normalized title. Returns (updated_records,
    saved_record, did_update)."""
    title_key = _normalize_record_key(record.get("title"))
    if title_key:
        for idx, existing in enumerate(records):
            if _normalize_record_key(existing.get("title")) == title_key:
                updated = dict(record)
                updated["sequence"] = existing.get("sequence") or idx + 1
                records[idx] = updated
                return records, updated, True

    next_record = dict(record)
    next_record["sequence"] = len(records) + 1
    records.append(next_record)
    return records, next_record, False


def append_record(record: dict[str, Any], path: str | Path = COMPLIANCE_LEDGER_JSON_PATH) -> list[dict[str, Any]]:
    records = load_records(path)
    records, _, _ = upsert_record(records, record)
    save_records(records, path)
    return records


def _normalize_procedure(value: Any) -> str:
    """Module-local copy of the procedure-normalizer that the workbook
    builder needs. compliance_ledger.py has its own (used during item
    normalisation); keeping the workbook layer self-contained avoids a
    circular import."""
    if not value:
        return "总办会审议"
    text = str(value).strip()
    if "董事" in text:
        return "董事会审议"
    if "总办" in text or "总经理" in text or "经理办" in text:
        return "总办会审议"
    return text or "总办会审议"


def create_compliance_workbook(
    records: list[dict[str, Any]],
    output_path: str | Path = COMPLIANCE_LEDGER_EXCEL_PATH,
) -> None:
    """Render the compliance ledger as an Excel workbook with merged
    per-record cells and a per-review row breakdown."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = WORKSHEET_NAME

    ws.merge_cells("A1:J1")
    ws["A1"] = "合规管理牵头部门合规审查工作台账"
    ws["A1"].font = Font(name="宋体", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 46.95

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(2, col, header)
        cell.font = Font(name="宋体", bold=True, size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 40.8

    widths = [13, 28, 16, 14, 28, 19, 32, 20, 16, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    row_idx = 3
    for idx, record in enumerate(records, start=1):
        review_rows = record.get("review_rows") or []
        if not review_rows:
            review_rows = [{
                "review_time": "", "review_unit": "",
                "review_opinion": "同意", "detail": "/", "implementation": "/",
            }]
        start = row_idx
        end = row_idx + len(review_rows) - 1

        merged_values = {
            1: record.get("sequence") or idx,
            2: record.get("title") or "",
            3: _normalize_procedure(record.get("procedure")),
            9: record.get("undertaking_department") or "法务合规部",
            10: "、".join(record.get("background_materials") or []) or "/",
        }
        for col, value in merged_values.items():
            ws.cell(start, col, value)
            if end > start:
                ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)

        for offset, review in enumerate(review_rows):
            r = start + offset
            ws.cell(r, 4, review.get("review_time") or "")
            ws.cell(r, 5, review.get("review_unit") or "")
            ws.cell(r, 6, review.get("review_opinion") or "同意")
            ws.cell(r, 7, review.get("detail") or "/")
            ws.cell(r, 8, review.get("implementation") or "/")
            ws.row_dimensions[r].height = 36

        for r in range(start, end + 1):
            for c in range(1, 11):
                cell = ws.cell(r, c)
                cell.font = Font(name="宋体", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
        row_idx = end + 1

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
