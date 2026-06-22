from __future__ import annotations

import io
import json
import re
import secrets
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from config import DATA_ROOT
from file_store import atomic_write_text, safe_child_path


PENDING_IMPORT_ROOT = Path(DATA_ROOT) / "_pending_ledger_imports"
IMPORT_TOKEN_PREFIX = "import_"


def normalize_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def cell_to_jsonable(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return value


def clean_text(value: Any) -> str:
    value = cell_to_jsonable(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, int | float):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group()) if match else None


def parse_int(value: Any) -> int:
    parsed = parse_float(value)
    return int(parsed) if parsed is not None else 0


def split_items(value: Any) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return [item.strip() for item in re.split(r"[、,，;；\n]+", text) if item.strip()]


def load_xlsx(file_bytes: bytes):
    try:
        return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError("无法读取 Excel，请确认上传的是 .xlsx 文件且未损坏。") from exc


def find_header_row(ws, required_headers: list[str], max_scan_rows: int = 10) -> int:
    best_row = 1
    best_score = -1
    required = [normalize_key(h) for h in required_headers]
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [normalize_key(cell.value) for cell in ws[row_idx]]
        score = sum(1 for wanted in required if wanted in values)
        non_empty = sum(1 for value in values if value)
        weighted = score * 100 + non_empty
        if weighted > best_score:
            best_row = row_idx
            best_score = weighted
    return best_row


def header_map(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        key = clean_text(cell.value)
        if key:
            mapping[key] = idx
    return mapping


def get_by_alias(row: dict[str, Any], aliases: list[str]) -> Any:
    normalized = {normalize_key(k): v for k, v in row.items()}
    for alias in aliases:
        key = normalize_key(alias)
        if key in normalized:
            return normalized[key]
    return None


def iter_rows_by_header(ws, required_headers: list[str], max_scan_rows: int = 10) -> list[dict[str, Any]]:
    header_row = find_header_row(ws, required_headers, max_scan_rows=max_scan_rows)
    headers = header_map(ws, header_row)
    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        item: dict[str, Any] = {}
        has_value = False
        for header, col_idx in headers.items():
            value = ws.cell(row_idx, col_idx).value
            if value not in (None, ""):
                has_value = True
            item[header] = cell_to_jsonable(value)
        if has_value:
            rows.append(item)
    return rows


def _merged_value(ws, row: int, col: int) -> Any:
    value = ws.cell(row, col).value
    if value is not None:
        return value
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(merged.min_row, merged.min_col).value
    return None


def _row_value(ws, row: int, headers: dict[str, int], aliases: list[str], *, merged: bool = False) -> Any:
    normalized = {normalize_key(k): v for k, v in headers.items()}
    for alias in aliases:
        col_idx = normalized.get(normalize_key(alias))
        if col_idx:
            value = _merged_value(ws, row, col_idx) if merged else ws.cell(row, col_idx).value
            return cell_to_jsonable(value)
    return None


def make_preview(
    *,
    records: list[dict[str, Any]],
    invalid_rows: list[dict[str, Any]],
    existing_keys: set[str],
    key_fn,
    import_token: str,
) -> dict[str, Any]:
    inserts = 0
    updates = 0
    for record in records:
        key = key_fn(record)
        if key and key in existing_keys:
            updates += 1
        else:
            inserts += 1
    return {
        "import_token": import_token,
        "rows_total": len(records) + len(invalid_rows),
        "rows_valid": len(records),
        "rows_invalid": len(invalid_rows),
        "inserts": inserts,
        "updates": updates,
        "invalid_rows": invalid_rows[:50],
        "sample_rows": records[:8],
    }


def _pending_kind_dir(user_id: int, kind: str) -> Path:
    safe_kind = re.sub(r"[^a-zA-Z0-9_-]", "_", kind)
    return PENDING_IMPORT_ROOT / safe_kind / f"user_{user_id}"


def create_pending_import(user_id: int, kind: str, payload: dict[str, Any]) -> str:
    token = f"{IMPORT_TOKEN_PREFIX}{secrets.token_urlsafe(16)}"
    target_dir = _pending_kind_dir(user_id, kind)
    target_dir.mkdir(parents=True, exist_ok=True)
    path = safe_child_path(target_dir, f"{token}.json")
    atomic_write_text(
        path,
        json.dumps(
            {
                "kind": kind,
                "created_at": datetime.now().isoformat(timespec="seconds"),
                **payload,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return token


def load_pending_import(user_id: int, kind: str, token: str) -> dict[str, Any]:
    if not token.startswith(IMPORT_TOKEN_PREFIX) or not token.replace("_", "").replace("-", "").isalnum():
        raise ValueError("无效的导入编号")
    path = safe_child_path(_pending_kind_dir(user_id, kind), f"{token}.json")
    if not path.exists():
        raise ValueError("导入预览已过期，请重新上传。")
    data = json.loads(path.read_text(encoding="utf-8"))
    if data.get("kind") != kind:
        raise ValueError("导入类型不匹配")
    return data


def parse_training_rows(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_xlsx(file_bytes)
    ws = wb.active
    raw_rows = iter_rows_by_header(ws, ["培训日期", "培训主题"])
    records: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []
    for idx, row in enumerate(raw_rows, start=1):
        record = {
            "date": clean_text(get_by_alias(row, ["培训日期", "日期"])),
            "topic": clean_text(get_by_alias(row, ["培训主题", "主题"])),
            "location": clean_text(get_by_alias(row, ["培训地点", "地点"])),
            "department": clean_text(get_by_alias(row, ["主办部门", "部门"])),
            "count": parse_int(get_by_alias(row, ["参与人数", "人数"])),
            "duration_hours": parse_float(get_by_alias(row, ["培训时长（课时）", "培训时长", "课时"])) or 0.0,
            "category": clean_text(get_by_alias(row, ["培训类别", "类别"])),
            "archive_path": clean_text(get_by_alias(row, ["归档路径", "附件路径"])),
        }
        missing = [name for name in ["date", "topic"] if not record[name]]
        if missing:
            invalid.append({"row": idx, "reason": f"缺少必填字段：{', '.join(missing)}", "data": row})
        else:
            records.append(record)
    return records, invalid


def parse_auth_rows(file_bytes: bytes, headers: list[str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_xlsx(file_bytes)
    ws = wb.active
    raw_rows = iter_rows_by_header(ws, ["编号", "题目"])
    records: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []
    for idx, row in enumerate(raw_rows, start=1):
        record = {header: cell_to_jsonable(get_by_alias(row, [header])) for header in headers}
        if not clean_text(record.get("编号")) and not clean_text(record.get("题目")):
            invalid.append({"row": idx, "reason": "缺少编号和题目，无法识别授权记录", "data": row})
        else:
            records.append(record)
    return records, invalid


def parse_legal_case_rows(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_xlsx(file_bytes)
    ws = wb.active
    header_row = find_header_row(ws, ["案件名称", "案由", "诉讼主体"], max_scan_rows=8)
    headers = header_map(ws, header_row)
    result_col = headers.get("处理结果")
    if result_col and "审级" not in headers and result_col < ws.max_column:
        headers["审级"] = result_col + 1
    records_by_key: dict[str, dict[str, Any]] = {}
    invalid: list[dict[str, Any]] = []
    current_key = ""

    for row_idx in range(header_row + 1, ws.max_row + 1):
        if all(ws.cell(row_idx, col).value in (None, "") for col in range(1, ws.max_column + 1)):
            continue
        seq = clean_text(_row_value(ws, row_idx, headers, ["序号"], merged=True))
        case_name = clean_text(_row_value(ws, row_idx, headers, ["案件名称"], merged=True))
        case_key = seq or normalize_key(case_name)
        if case_key:
            current_key = case_key
        if not current_key:
            invalid.append({"row": row_idx, "reason": "缺少案件名称或序号，无法归组", "data": {}})
            continue

        record = records_by_key.get(current_key)
        if record is None:
            record = {
                "案件名称": case_name,
                "案件发生时间": clean_text(_row_value(ws, row_idx, headers, ["案件发生时间"], merged=True)) or None,
                "案由": clean_text(_row_value(ws, row_idx, headers, ["案由"], merged=True)),
                "诉讼主体": clean_text(_row_value(ws, row_idx, headers, ["诉讼主体"], merged=True)),
                "主诉被诉": clean_text(_row_value(ws, row_idx, headers, ["主诉/被诉", "主诉被诉"], merged=True)),
                "标的金额": parse_float(_row_value(ws, row_idx, headers, ["标的金额（万元）", "标的金额"], merged=True)),
                "基本情况": clean_text(_row_value(ws, row_idx, headers, ["基本情况（来源于业务单位情况说明及起诉状诉讼请求）", "基本情况"], merged=True)),
                "生效判决日期": clean_text(_row_value(ws, row_idx, headers, ["生效判决/裁定日期", "生效判决日期"], merged=True)) or None,
                "强制执行时间": clean_text(_row_value(ws, row_idx, headers, ["强制执行时间"], merged=True)) or None,
                "服务律所": clean_text(_row_value(ws, row_idx, headers, ["服务律所（人工填写项）", "服务律所"], merged=True)) or None,
                "stages": [],
                "案号列表": [],
            }
            if not record["案件名称"]:
                invalid.append({"row": row_idx, "reason": "缺少案件名称", "data": record})
                continue
            records_by_key[current_key] = record

        stage_name = clean_text(_row_value(ws, row_idx, headers, ["审级"], merged=False))
        result = clean_text(_row_value(ws, row_idx, headers, ["处理结果"], merged=False))
        if stage_name or result:
            record.setdefault("stages", []).append({"审级": stage_name, "处理结果": result})

    return list(records_by_key.values()), invalid


def parse_compliance_rows(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_xlsx(file_bytes)
    ws = wb.active
    header_row = find_header_row(ws, ["重大事项", "审查单位", "合规审查意见"], max_scan_rows=8)
    headers = header_map(ws, header_row)
    records_by_key: dict[str, dict[str, Any]] = {}
    invalid: list[dict[str, Any]] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        if all(ws.cell(row_idx, col).value in (None, "") for col in range(1, ws.max_column + 1)):
            continue
        seq = clean_text(_row_value(ws, row_idx, headers, ["序号"], merged=True))
        title = clean_text(_row_value(ws, row_idx, headers, ["重大事项"], merged=True))
        key = seq or normalize_key(title)
        if not key or not title:
            invalid.append({"row": row_idx, "reason": "缺少重大事项标题", "data": {}})
            continue

        record = records_by_key.get(key)
        if record is None:
            record = {
                "title": title,
                "procedure": clean_text(_row_value(ws, row_idx, headers, ["程序"], merged=True)),
                "undertaking_department": clean_text(_row_value(ws, row_idx, headers, ["承办单位"], merged=True)) or "法务合规部",
                "background_materials": split_items(_row_value(ws, row_idx, headers, ["背景材料"], merged=True)),
                "review_rows": [],
                "warnings": [],
            }
            records_by_key[key] = record

        review = {
            "review_time": clean_text(_row_value(ws, row_idx, headers, ["审查时间"], merged=False)),
            "review_unit": clean_text(_row_value(ws, row_idx, headers, ["审查单位"], merged=False)),
            "review_opinion": clean_text(_row_value(ws, row_idx, headers, ["合规审查意见"], merged=False)) or "同意",
            "detail": clean_text(_row_value(ws, row_idx, headers, ["具体意见描述"], merged=False)) or "/",
            "implementation": clean_text(_row_value(ws, row_idx, headers, ["落实情况"], merged=False)) or "/",
        }
        if any(review.values()):
            record["review_rows"].append(review)

    return list(records_by_key.values()), invalid
