import os
import re
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from config import DATA_ROOT
from file_store import atomic_save_workbook, file_lock


EXCEL_PATH = os.path.join(DATA_ROOT, "培训统计表.xlsx")

HEADERS = ["序号", "培训日期", "培训主题", "培训地点", "主办部门", "参与人数", "培训时长（课时）", "培训类别", "归档路径", "录入时间"]
COL_WIDTHS = [6, 14, 30, 20, 16, 10, 14, 14, 50, 20]


def init_excel():
    """初始化 Excel 文件（如果不存在则创建）"""
    with file_lock(EXCEL_PATH):
        if os.path.exists(EXCEL_PATH):
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "培训记录"

        # 写表头
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 设置列宽
        for col, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 22
        atomic_save_workbook(wb, EXCEL_PATH)


def _migrate_headers(ws) -> None:
    """
    检测旧版表头，若缺少「培训时长（课时）」列则在「参与人数」之后插入，
    同时将已有数据行对应列右移，保持数据对齐。
    """
    existing = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if "培训时长（课时）" in existing:
        return  # 已是新版，无需迁移

    # 找到「参与人数」所在列（1-based），新列插在其后
    insert_after = next((i + 1 for i, h in enumerate(existing) if h == "参与人数"), None)
    if insert_after is None:
        return  # 表头格式未知，跳过

    insert_col = insert_after + 1  # 新列的列号

    # openpyxl insert_cols 在指定列前插入
    ws.insert_cols(insert_col)

    # 写新列表头样式
    cell = ws.cell(row=1, column=insert_col, value="培训时长（课时）")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置新列列宽
    ws.column_dimensions[openpyxl.utils.get_column_letter(insert_col)].width = 14

    # 已有数据行在新列填空（insert_cols 已自动移位，新列单元格为空，无需额外处理）


def _normalize_key(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _find_existing_training_row(ws, date: str, topic: str) -> int | None:
    target_date = _normalize_key(date)
    target_topic = _normalize_key(topic)
    if not target_date or not target_topic:
        return None

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    date_col = headers.get("培训日期")
    topic_col = headers.get("培训主题")
    if not date_col or not topic_col:
        return None

    for row in range(2, ws.max_row + 1):
        if (
            _normalize_key(ws.cell(row=row, column=date_col).value) == target_date
            and _normalize_key(ws.cell(row=row, column=topic_col).value) == target_topic
        ):
            return row
    return None


def append_record(
    date: str,
    topic: str,
    location: str,
    department: str,
    count: int,
    category: str,
    archive_path: str,
    duration_hours: float = 0.0,
):
    """写入培训记录；同一培训日期和主题已存在时更新原行。"""
    with file_lock(EXCEL_PATH):
        if os.path.exists(EXCEL_PATH):
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            _migrate_headers(ws)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "培训记录"
            for col, header in enumerate(HEADERS, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col, width in enumerate(COL_WIDTHS, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            ws.row_dimensions[1].height = 22

        existing_row = _find_existing_training_row(ws, date, topic)
        next_row = existing_row or ws.max_row + 1
        seq = ws.cell(row=next_row, column=1).value if existing_row else next_row - 1
        seq = seq or next_row - 1

        now = datetime.now().strftime("%Y-%m-%d %H:%M")

        values = [seq, date, topic, location, department, count, duration_hours or "", category, archive_path, now]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.alignment = Alignment(vertical="center")
            # 奇偶行背景色
            if next_row % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", fgColor="DCE6F1")

        atomic_save_workbook(wb, EXCEL_PATH)
    return EXCEL_PATH
