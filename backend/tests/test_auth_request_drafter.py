import sys
import tempfile
import unittest
from pathlib import Path


BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))


class AuthRequestDrafterTests(unittest.TestCase):
    def test_attachment1_title_uses_document_heading(self):
        from utils.auth_request_drafter import _extract_attachment1_title

        text = "\n".join(
            [
                "— 1 —",
                "关于国航T3 北区机务设施建设项目",
                "需求调整及明确相关工作的通知",
                "中国国际航空股份有限公司工程建设管理部：",
                "《关于办理调整T3北区项目Ameco需求的请示》我部已收悉。",
            ]
        )

        self.assertEqual(
            _extract_attachment1_title(text),
            "关于国航T3北区机务设施建设项目需求调整及明确相关工作的通知",
        )

    def test_build_request_text_falls_back_from_polluted_extraction(self):
        from utils.auth_request_drafter import build_auth_request_text

        polluted = "关于办理调整T3北区项目Ameco需求的请示" * 10
        content = build_auth_request_text(
            {
                "attachment1": {
                    "title": polluted,
                    "project_name": polluted,
                    "document_no": "国航股份有限公司中航集团规划发〔2026〕1号",
                    "undertaking_unit": polluted,
                },
                "attachment2": {
                    "principal_unit": "中国航空集团建设开发有限公司",
                    "principal_short": "建开公司",
                    "trustee_work_unit": "中国航空集团建设开发有限公司",
                    "trustee_position": "项目经理",
                    "trustee_name": "白利业白利业",
                    "permission_detail": "合同签署权限。合同签署权限。",
                    "permission_type": "合同签署权限",
                    "authorization_scope": "为开展国航T3北区机务设施建设项目的相关工作，其权限为:合同签署权限。",
                },
            },
            {"auth_mode": "direct", "copies": "3", "seal": "公章", "handler": "张三"},
        )

        self.assertIn("依据《关于国航T3北区机务设施建设项目需求调整及明确相关工作的通知》", content)
        self.assertIn("（中航集团规划发〔2026〕1号，详见附件1）", content)
        self.assertIn("白利业同志", content)
        self.assertNotIn("白利业白利业同志", content)

    def test_doc_binary_fallback_extracts_legacy_doc_text(self):
        from utils.auth_request_drafter import extract_attachment2_info

        text = (
            "\x00" * 32
            + "授权 委 托 书\r"
            + "委托单位\x07企业名称\x07中国航空集团建设开发有限公司\x07注册地址\x07北京市\x07"
            + "法定代表人\x07布赫\x07职务\x07董事长\x07"
            + "受托人\x07姓名\x07白利业\x07电话\x0713601268681\x07工作单位\x07中国航空集团建设开发有限公司工程建设管理中心\x07职务\x07总经理\x07"
            + "委托事项及权限\x07为开展国航T3北区机务设施建设项目的相关工作，其权限为:代表委托单位签署本项目相关合同。"
            + "\x07授权期限\x07授权期限：自2026年2月2日起至国航T3北区机务设施建设项目结束止。"
            + "\x07委托单位盖章\x07备注说明\x07一式3份\x07授权编号：建开转托字（2026）007号"
        )
        info = extract_attachment2_info(text.encode("utf-16le"), "legacy.doc")

        self.assertEqual(info["authorization_no"], "建开转托字（2026）007号")
        self.assertEqual(info["trustee_name"], "白利业")
        self.assertEqual(info["authorization_term"], "自2026年2月2日起至国航T3北区机务设施建设项目结束止。")

    def test_docx_uses_fang_song_for_latin_and_digits(self):
        from docx import Document
        from docx.oxml.ns import qn
        from utils.auth_request_drafter import save_auth_request_docx

        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            path = tmp.name
        try:
            save_auth_request_docx("ABC123测试", path)
            doc = Document(path)
            fonts = doc.paragraphs[0].runs[0]._r.rPr.rFonts
            self.assertEqual(fonts.get(qn("w:ascii")), "仿宋_GB2312")
            self.assertEqual(fonts.get(qn("w:hAnsi")), "仿宋_GB2312")
            self.assertEqual(fonts.get(qn("w:eastAsia")), "仿宋_GB2312")
        finally:
            Path(path).unlink(missing_ok=True)

    def test_record_to_ledger_updates_existing_authorization_no(self):
        import openpyxl
        from utils.auth_request_drafter import record_to_ledger

        extracted = {
            "attachment2": {
                "authorization_no": "建开转托字（2026）007号",
                "legal_representative": "布赫",
                "trustee_name": "白利业",
                "authorization_term": "授权期限：自2026年2月2日起至项目结束止。",
                "authorization_scope": "授权内容A",
            }
        }
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmpdir:
            ledger_path = str(Path(tmpdir) / "ledger.xlsx")
            record_to_ledger(extracted, {"handler": "张三", "seal": "公章", "copies": "3"}, "旧标题", ledger_path)
            record_to_ledger(
                {**extracted, "attachment2": {**extracted["attachment2"], "authorization_scope": "授权内容B"}},
                {"handler": "李四", "seal": "合同专用章", "copies": "5"},
                "新标题",
                ledger_path,
            )
            ws = openpyxl.load_workbook(ledger_path).active

            self.assertEqual(ws.max_row, 2)
            self.assertEqual(ws.cell(2, 1).value, 1)
            self.assertEqual(ws.cell(2, 3).value, "李四")
            self.assertEqual(ws.cell(2, 6).value, "合同专用章")
            self.assertEqual(ws.cell(2, 8).value, "自2026年2月2日起至项目结束止。")
            self.assertEqual(ws.cell(2, 9).value, "授权内容B")
            self.assertEqual(ws.cell(2, 14).value, "新标题")


if __name__ == "__main__":
    unittest.main()
