import io
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))


def workbook_bytes(headers, rows) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class LedgerImporterTests(unittest.TestCase):
    def test_parse_training_rows(self):
        from utils.ledger_importer import parse_training_rows

        data = workbook_bytes(
            ["序号", "培训日期", "培训主题", "培训地点", "主办部门", "参与人数", "培训时长（课时）", "培训类别", "归档路径"],
            [[1, "2026-01-02", "合规培训", "总部", "法务部", "23人", 2.5, "合规", "archive/path"]],
        )
        records, invalid = parse_training_rows(data)

        self.assertEqual(invalid, [])
        self.assertEqual(records[0]["topic"], "合规培训")
        self.assertEqual(records[0]["count"], 23)
        self.assertEqual(records[0]["duration_hours"], 2.5)

    def test_parse_auth_rows(self):
        from utils.auth_request_drafter import AUTH_LEDGER_HEADERS
        from utils.ledger_importer import parse_auth_rows

        data = workbook_bytes(
            AUTH_LEDGER_HEADERS,
            [[1, "AUTH-001", "张三", "授权人", "代理人", "公章", 2, "2026.1-2026.2", "处理事项", "", "", "", "", "测试授权", "", ""]],
        )
        records, invalid = parse_auth_rows(data, AUTH_LEDGER_HEADERS)

        self.assertEqual(invalid, [])
        self.assertEqual(records[0]["编号"], "AUTH-001")
        self.assertEqual(records[0]["题目"], "测试授权")

    def test_import_auth_ledger_rows_upserts_by_auth_no(self):
        from utils.auth_request_drafter import import_auth_ledger_rows

        with tempfile.TemporaryDirectory() as tmpdir:
            path = str(Path(tmpdir) / "授权委托台账.xlsx")
            first = [{"编号": "AUTH-001", "经办人": "张三", "题目": "旧题目"}]
            second = [{"编号": "AUTH-001", "经办人": "李四", "题目": "新题目"}]

            self.assertEqual(import_auth_ledger_rows(first, path), {"inserts": 1, "updates": 0, "count": 1})
            self.assertEqual(import_auth_ledger_rows(second, path), {"inserts": 0, "updates": 1, "count": 1})

            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            self.assertEqual(ws.max_row, 2)
            self.assertEqual(ws.cell(2, 3).value, "李四")
            self.assertEqual(ws.cell(2, 14).value, "新题目")

    def test_parse_legal_case_rows_from_generated_ledger(self):
        from utils.ledger_importer import parse_legal_case_rows
        from utils.write_excel import write_ledger

        case = {
            "案件名称": "测试案件",
            "案件发生时间": "2025-01-01",
            "案由": "合同纠纷",
            "诉讼主体": "原告：A\n被告：B",
            "主诉被诉": "主诉",
            "标的金额": 10.5,
            "基本情况": "基本情况",
            "生效判决日期": "2025-02-01",
            "强制执行时间": None,
            "服务律所": "测试律所",
            "stages": [
                {"审级": "一审", "处理结果": "一审结果"},
                {"审级": "二审", "处理结果": "二审结果"},
            ],
            "案号列表": [],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "cases.xlsx"
            write_ledger([case], str(path))
            records, invalid = parse_legal_case_rows(path.read_bytes())

        self.assertEqual(invalid, [])
        self.assertEqual(records[0]["案件名称"], "测试案件")
        self.assertEqual(records[0]["标的金额"], 10.5)
        self.assertEqual(len(records[0]["stages"]), 2)
        self.assertEqual(records[0]["stages"][0]["审级"], "一审")

    def test_parse_compliance_rows_from_generated_workbook(self):
        from utils.compliance_ledger import create_compliance_workbook
        from utils.ledger_importer import parse_compliance_rows

        record = {
            "sequence": 1,
            "title": "重大事项",
            "procedure": "总办会审议",
            "undertaking_department": "法务合规部",
            "background_materials": ["背景材料"],
            "review_rows": [
                {
                    "review_time": "2026-01-01",
                    "review_unit": "会签单位",
                    "review_opinion": "同意",
                    "detail": "/",
                    "implementation": "/",
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "compliance.xlsx"
            create_compliance_workbook([record], path)
            records, invalid = parse_compliance_rows(path.read_bytes())

        self.assertEqual(invalid, [])
        self.assertEqual(records[0]["title"], "重大事项")
        self.assertEqual(records[0]["review_rows"][0]["review_unit"], "会签单位")


if __name__ == "__main__":
    unittest.main()
