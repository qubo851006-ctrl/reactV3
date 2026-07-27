import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parents[1]
TEST_TMP_ROOT = BACKEND_DIR / "tests" / "tmp"
sys.path.insert(0, str(BACKEND_DIR))

from utils.compliance_ledger import (  # noqa: E402
    _deduplicate_chief_opinion,
    _detail_for_opinion,
    _fix_chief_opinion_from_text,
    _is_compliance_department,
    append_record,
    build_review_rows,
    create_compliance_workbook,
    extract_compliance_item,
    load_responsible_persons,
    normalize_extracted_item,
    normalize_review_opinion,
    save_responsible_persons,
)


class _FakeMessage:
    def __init__(self, content: str):
        self.content = content


class _FakeChoice:
    def __init__(self, content: str):
        self.message = _FakeMessage(content)


class _FakeResponse:
    def __init__(self, content: str):
        self.choices = [_FakeChoice(content)]


class _FakeCompletions:
    def __init__(self, responses: list[str]):
        self.responses = list(responses)
        self.models: list[str] = []

    def create(self, **kwargs):
        self.models.append(kwargs["model"])
        return _FakeResponse(self.responses.pop(0))


class _FakeChat:
    def __init__(self, completions: _FakeCompletions):
        self.completions = completions


class _FakeClient:
    def __init__(self, completions: _FakeCompletions):
        self.chat = _FakeChat(completions)


class ComplianceLedgerOpinionTests(unittest.TestCase):
    def test_normalize_review_opinion_prefers_disagreement(self):
        self.assertEqual(normalize_review_opinion("不同意该方案，建议重新论证"), "不予同意")

    def test_normalize_review_opinion_detects_supplement_suggestion(self):
        self.assertEqual(normalize_review_opinion("拟同意，建议补充预算测算依据"), "建议补充完善")

    def test_normalize_review_opinion_defaults_to_agree(self):
        self.assertEqual(normalize_review_opinion("拟同意。"), "同意")

    def test_normalize_review_opinion_treats_meeting_submission_suggestion_as_agree(self):
        self.assertEqual(normalize_review_opinion("拟同意，建议提交总经理办公会议审议。"), "同意")


class ComplianceLedgerRowsTests(unittest.TestCase):
    def test_build_review_rows_expands_multiple_countersign_departments(self):
        item = {
            "undertaking": {
                "department": "财务部",
                "person": "杨焕",
                "time": "2026-05-01",
                "opinion_text": "拟同意。",
            },
            "countersign": [
                {
                    "department": "人力资源部",
                    "person": "陈锐",
                    "time": "2026-05-02",
                    "opinion_text": "建议补充人员安排。",
                    "detail": "建议补充人员安排。",
                    "implementation": "已按要求补充完善",
                },
                {
                    "department": "财务部",
                    "person": "杨焕",
                    "time": "2026-05-03",
                    "opinion_text": "同意。",
                },
            ],
            "compliance": {
                "department": "审计部/法务合规部",
                "person": "李莹",
                "time": "2026-05-04",
                "opinion_text": "拟同意。",
            },
            "chief": {
                "person": "胡鹏斌",
                "time": "2026-05-05",
                "opinion_text": "拟同意。",
            },
        }

        rows = build_review_rows(item)

        self.assertEqual([row["review_unit"] for row in rows], [
            "首席合规官",
            "合规管理牵头部门（审计部/法务合规部）",
            "会签单位（人力资源部）",
            "会签单位（财务部）",
            "承办单位（财务部）",
        ])
        self.assertEqual(rows[2]["review_opinion"], "建议补充完善")
        self.assertEqual(rows[2]["detail"], "建议补充人员安排。")
        self.assertEqual(rows[2]["implementation"], "已按要求补充完善")

    def test_approval_entries_assign_chief_by_signer_without_merging_neighbor_opinion(self):
        item = normalize_extracted_item({
            "title": "关于测试事项的请示",
            "procedure": "总办会审议",
            "approval_entries": [
                {
                    "department": "中航建设直属",
                    "person": "徐勤",
                    "time": "2026-05-11 10:40:42",
                    "opinion_text": "同意提交总办会审议。请履行会前传签程序。",
                },
                {
                    "department": "中航建设直属",
                    "person": "胡鹏斌",
                    "time": "2026-05-11 10:21:13",
                    "opinion_text": "拟同意，建议提交总经理办公会议审议。",
                },
            ],
            "warnings": [],
        })

        rows = item["review_rows"]
        chief_rows = [row for row in rows if row["review_unit"] == "首席合规官"]

        self.assertEqual(len(chief_rows), 1)
        self.assertEqual(chief_rows[0]["review_time"], "2026-05-11 10:21:13")
        self.assertEqual(chief_rows[0]["review_opinion"], "同意")
        self.assertEqual(chief_rows[0]["detail"], "/")
        self.assertNotIn("徐勤", json.dumps(chief_rows[0], ensure_ascii=False))
        self.assertNotIn("请履行会前传签程序", json.dumps(chief_rows[0], ensure_ascii=False))
        self.assertNotIn("会签单位（中航建设直属）", [row["review_unit"] for row in rows])

    def test_approval_entries_pick_configured_heads_and_ignore_unknown_persons(self):
        item = normalize_extracted_item({
            "title": "关于测试事项的请示",
            "procedure": "总办会审议",
            "approval_entries": [
                {
                    "department": "中航建设直属",
                    "person": f"经办人{i}",
                    "time": f"2026-05-11 {10 + i}:00:00",
                    "opinion_text": "同意。",
                }
                for i in range(6)
            ] + [
                {
                    "department": "人力资源部",
                    "person": "陈锐",
                    "time": "2026-05-11 17:00:00",
                    "opinion_text": "同意。",
                }
            ],
        })

        review_units = [row["review_unit"] for row in item["review_rows"]]

        self.assertIn("会签单位（人力资源部）", review_units)
        self.assertNotIn("会签单位（中航建设直属）", review_units)

    def test_countersign_merges_explicit_field_and_approval_entries(self):
        item = normalize_extracted_item({
            "title": "关于测试事项的请示",
            "procedure": "总办会审议",
            "approval_entries": [
                {
                    "department": "人力资源部",
                    "person": "陈锐",
                    "time": "2026-05-11 17:00:00",
                    "opinion_text": "同意。",
                },
            ],
            "countersign": [
                {
                    "department": "财务部",
                    "person": "杨焕",
                    "time": "2026-05-11 18:00:00",
                    "opinion_text": "同意。",
                },
            ],
        })

        review_units = [row["review_unit"] for row in item["review_rows"]]

        self.assertIn("会签单位（财务部）", review_units)
        self.assertIn("会签单位（人力资源部）", review_units)

    def test_countersign_filters_to_configured_department_heads(self):
        responsible_persons = {
            "财务部": "杨焕",
            "西南分公司（四川中航物业）": "张虎",
            "党群办公室/董事会办公室/行政办公室": "刘芳",
        }
        item = normalize_extracted_item({
            "title": "关于测试事项的请示",
            "procedure": "总办会审议",
            "countersign": [
                {
                    "department": "西南分公司（四川中航物业） 直属",
                    "person": "张虎",
                    "time": "2026-05-09 09:53:27",
                    "opinion_text": "已阅。",
                },
                {
                    "department": "西南分公司（四川中航物业）",
                    "person": "王永君",
                    "time": "2026-05-09 09:35:13",
                    "opinion_text": "已核。",
                },
                {
                    "department": "财务部",
                    "person": "杨焕",
                    "time": "2026-05-09 15:08:46",
                    "opinion_text": "拟同意。",
                },
            ],
        }, responsible_persons)

        review_units = [row["review_unit"] for row in item["review_rows"]]

        self.assertEqual(review_units, ["会签单位（西南分公司（四川中航物业））", "会签单位（财务部）"])

    def test_approval_entries_fill_countersign_for_all_configured_heads(self):
        responsible_persons = {
            "人力资源部": "陈锐",
            "党群办公室/董事会办公室/行政办公室": "刘芳",
        }
        item = normalize_extracted_item({
            "title": "关于测试事项的请示",
            "procedure": "总办会审议",
            "approval_entries": [
                {
                    "department": "党群办公室/董事会办公室/行政办公室",
                    "person": "刘芳",
                    "time": "2026-05-11 09:49:20",
                    "opinion_text": "已阅。",
                },
                {
                    "department": "人力资源部",
                    "person": "陈锐",
                    "time": "2026-05-11 09:50:00",
                    "opinion_text": "同意。",
                },
            ],
        }, responsible_persons)

        review_units = [row["review_unit"] for row in item["review_rows"]]

        self.assertIn("会签单位（党群办公室/董事会办公室/行政办公室）", review_units)
        self.assertIn("会签单位（人力资源部）", review_units)


class ComplianceLedgerWorkbookTests(unittest.TestCase):
    def test_create_workbook_merges_matter_columns_and_expands_review_rows(self):
        TEST_TMP_ROOT.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmpdir:
            output = Path(tmpdir) / "ledger.xlsx"
            records = [{
                "sequence": 1,
                "title": "关于计划外新增企业知识库建设项目（一期）及立项的请示",
                "procedure": "总办会审议",
                "undertaking_department": "法务合规部",
                "background_materials": ["项目立项报告", "预算测算表"],
                "review_rows": [
                    {
                        "review_time": "2026-05-05",
                        "review_unit": "首席合规官",
                        "review_opinion": "同意",
                        "detail": "/",
                        "implementation": "/",
                    },
                    {
                        "review_time": "2026-05-04",
                        "review_unit": "合规管理牵头部门（审计部/法务合规部）",
                        "review_opinion": "建议补充完善",
                        "detail": "建议补充预算测算依据",
                        "implementation": "已按要求补充完善",
                    },
                    {
                        "review_time": "2026-05-03",
                        "review_unit": "会签单位（财务部）",
                        "review_opinion": "同意",
                        "detail": "/",
                        "implementation": "/",
                    },
                    {
                        "review_time": "2026-05-01",
                        "review_unit": "承办单位（财务部）",
                        "review_opinion": "同意",
                        "detail": "/",
                        "implementation": "/",
                    },
                ],
            }]

            create_compliance_workbook(records, output)
            wb = openpyxl.load_workbook(output)
            ws = wb["合规管理牵头部门合规审查台账"]

            self.assertEqual(ws["A1"].value, "合规管理牵头部门合规审查工作台账")
            self.assertEqual(ws["A2"].value, "序号")
            self.assertEqual(ws["A3"].value, 1)
            self.assertEqual(ws["B3"].value, records[0]["title"])
            self.assertEqual(ws["C3"].value, "总办会审议")
            self.assertEqual(ws["E4"].value, "合规管理牵头部门（审计部/法务合规部）")
            self.assertEqual(ws["F4"].value, "建议补充完善")
            self.assertEqual(ws["J3"].value, "项目立项报告、预算测算表")
            self.assertIn("A3:A6", [str(rng) for rng in ws.merged_cells.ranges])
            self.assertIn("B3:B6", [str(rng) for rng in ws.merged_cells.ranges])


class ComplianceLedgerPersistenceTests(unittest.TestCase):
    def test_responsible_persons_config_can_be_saved_and_reloaded(self):
        TEST_TMP_ROOT.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmpdir:
            path = Path(tmpdir) / "responsible_persons.json"

            saved = save_responsible_persons({"财务部": "杨焕", " 空部门 ": " "}, path)
            loaded = load_responsible_persons(path)

            self.assertEqual(saved, {"财务部": "杨焕"})
            self.assertEqual(loaded["财务部"], "杨焕")
            self.assertNotIn(" 空部门 ", loaded)

    def test_append_record_assigns_natural_sequence_numbers(self):
        TEST_TMP_ROOT.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmpdir:
            path = Path(tmpdir) / "records.json"
            base_record = {
                "title": "事项",
                "procedure": "董事会审议",
                "undertaking_department": "法务合规部",
                "background_materials": [],
                "review_rows": [],
            }

            records = append_record(base_record, path)
            records = append_record({**base_record, "title": "事项二"}, path)

            self.assertEqual([r["sequence"] for r in records], [1, 2])
            self.assertEqual(records[1]["title"], "事项二")

    def test_append_record_updates_existing_item_by_title(self):
        TEST_TMP_ROOT.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmpdir:
            path = Path(tmpdir) / "records.json"
            base_record = {
                "title": "关于采购系统升级的合规审查",
                "procedure": "董事会审议",
                "undertaking_department": "法务合规部",
                "background_materials": ["旧附件.pdf"],
                "review_rows": [{"review_unit": "审计部/法务合规部", "detail": "旧意见"}],
            }

            records = append_record(base_record, path)
            records = append_record({
                **base_record,
                "procedure": "总办会审议",
                "background_materials": ["新附件.pdf"],
                "review_rows": [{"review_unit": "审计部/法务合规部", "detail": "新意见"}],
            }, path)

            self.assertEqual(len(records), 1)
            self.assertEqual(records[0]["sequence"], 1)
            self.assertEqual(records[0]["procedure"], "总办会审议")
            self.assertEqual(records[0]["background_materials"], ["新附件.pdf"])
            self.assertEqual(records[0]["review_rows"][0]["detail"], "新意见")

    def test_write_compliance_rolls_back_json_when_workbook_generation_fails(self):
        from routers.compliance import ComplianceWriteRequest, write_compliance

        TEST_TMP_ROOT.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmpdir:
            json_path = Path(tmpdir) / "records.json"
            excel_path = Path(tmpdir) / "ledger.xlsx"
            json_path.write_text(json.dumps([{
                "sequence": 1,
                "title": "旧事项",
                "procedure": "董事会审议",
                "undertaking_department": "法务合规部",
                "background_materials": [],
                "review_rows": [],
            }], ensure_ascii=False), encoding="utf-8")
            excel_path.write_bytes(b"old-excel")

            body = ComplianceWriteRequest(
                title="新事项",
                procedure="董事会审议",
                undertaking_department="法务合规部",
                background_materials=[],
                review_rows=[],
            )
            fake_db = type("DB", (), {"add": lambda *_: None, "commit": lambda *_: None})()
            fake_user = type("User", (), {"id": 1})()

            with patch("routers.compliance.COMPLIANCE_LEDGER_JSON_PATH", str(json_path)), \
                 patch("routers.compliance.COMPLIANCE_LEDGER_EXCEL_PATH", str(excel_path)), \
                 patch("routers.compliance.create_compliance_workbook", side_effect=RuntimeError("boom")), \
                 patch("routers.compliance.write_log"):
                with self.assertRaises(Exception):
                    write_compliance(body, request=None, db=fake_db, user=fake_user)

            records = json.loads(json_path.read_text(encoding="utf-8"))
            self.assertEqual([r["title"] for r in records], ["旧事项"])
            self.assertEqual(excel_path.read_bytes(), b"old-excel")


class ComplianceLedgerModelTests(unittest.TestCase):
    def test_extract_compliance_item_uses_qwen_then_deepseek_review(self):
        qwen_json = """
        {
          "title": "原始标题",
          "procedure": "董事会审议",
          "attachments": ["附件一.pdf"],
          "undertaking": {"department": "财务部", "person": "杨焕", "time": "2026-05-01", "opinion_text": "拟同意", "detail": "", "implementation": "/"},
          "countersign": [],
          "compliance": {"department": "审计部/法务合规部", "person": "李莹", "time": "2026-05-02", "opinion_text": "拟同意", "detail": "", "implementation": "/"},
          "chief": {"person": "胡鹏斌", "time": "2026-05-03", "opinion_text": "拟同意", "detail": "", "implementation": "/"},
          "warnings": []
        }
        """
        deepseek_json = """
        {
          "title": "校验后的标题",
          "procedure": "总办会审议",
          "attachments": ["附件二.xlsx"],
          "undertaking": {"department": "财务部", "person": "杨焕", "time": "2026-05-01", "opinion_text": "拟同意", "detail": "", "implementation": "/"},
          "countersign": [],
          "compliance": {"department": "审计部/法务合规部", "person": "李莹", "time": "2026-05-02", "opinion_text": "拟同意", "detail": "", "implementation": "/"},
          "chief": {"person": "胡鹏斌", "time": "2026-05-03", "opinion_text": "拟同意", "detail": "", "implementation": "/"},
          "warnings": ["DeepSeek 已校验"]
        }
        """
        completions = _FakeCompletions([qwen_json, deepseek_json])

        with patch("llm_client.get_llm_client", return_value=_FakeClient(completions)):
            item = extract_compliance_item("OA正文内容", {"财务部": "杨焕"})

        self.assertEqual(completions.models, ["qwen3.6", "deepseek-v4-flash"])
        self.assertNotIn("glm-5-outside", completions.models)
        self.assertEqual(item["title"], "校验后的标题")
        self.assertEqual(item["procedure"], "总办会审议")
        self.assertEqual(item["background_materials"], ["附件二"])
        self.assertIn("DeepSeek 已校验", item["warnings"])

    def test_extract_compliance_item_supplements_countersign_heads_from_source_text(self):
        qwen_json = """
        {
          "title": "关于测试事项的请示",
          "procedure": "总办会审议",
          "attachments": [],
          "undertaking": {"department": "规划与资产部/深化改革领导小组办公室", "person": "富小鹏", "time": "2026-05-08 18:35:46", "opinion_text": "拟同意。", "detail": "", "implementation": "/"},
          "countersign": [
            {"department": "财务部", "person": "杨焕", "time": "2026-05-09 15:08:46", "opinion_text": "拟同意。", "detail": "", "implementation": "/"}
          ],
          "compliance": {"department": "审计部/法务合规部", "person": "李莹", "time": "2026-05-09 10:51:09", "opinion_text": "已阅。拟同意。", "detail": "", "implementation": "/"},
          "chief": {"person": "胡鹏斌", "time": "2026-05-11 10:21:13", "opinion_text": "拟同意，建议提交总经理办公会会议审议。", "detail": "", "implementation": "/"},
          "warnings": []
        }
        """
        deepseek_json = qwen_json
        source_text = """
签发意见
拟同意，建议提交总经理办公会会议审议。
中航建设直属 胡鹏斌 2026-05-11 10:21:13
会签
拟同意。
财务部 杨焕 2026-05-09 15:08:46
已阅。
审计部/法务合规部 李莹 2026-05-09 10:51:09
已阅。
西南分公司（四川中航物业） 直属 张 虎 2026-05-09 09:53:27
已阅。
西南分公司（四川中航物业） 直属 王永君 2026-05-09 09:35:13
已阅。
党群办公室/董事会办公室/行政办公室 刘芳 2026-05-11 09:49:20
批准部门意见
"""
        responsible_persons = {
            "财务部": "杨焕",
            "审计部/法务合规部": "李莹",
            "西南分公司（四川中航物业）": "张虎",
            "党群办公室/董事会办公室/行政办公室": "刘芳",
            "规划与资产部/深化改革领导小组办公室": "富小鹏",
        }
        completions = _FakeCompletions([qwen_json, deepseek_json])

        with patch("llm_client.get_llm_client", return_value=_FakeClient(completions)):
            item = extract_compliance_item(source_text, responsible_persons)

        review_units = [row["review_unit"] for row in item["review_rows"]]

        self.assertIn("会签单位（财务部）", review_units)
        self.assertIn("会签单位（西南分公司（四川中航物业））", review_units)
        self.assertIn("会签单位（党群办公室/董事会办公室/行政办公室）", review_units)
        self.assertNotIn("会签单位（审计部/法务合规部）", review_units)
        self.assertNotIn("王永君", json.dumps(item["review_rows"], ensure_ascii=False))

    def test_extract_compliance_item_resets_chief_detail_from_source_text(self):
        merged_chief_text = "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会会议审议。"
        qwen_json = f"""
        {{
          "title": "关于测试事项的请示",
          "procedure": "总办会审议",
          "attachments": [],
          "undertaking": {{"department": "规划与资产部/深化改革领导小组办公室", "person": "富小鹏", "time": "2026-05-08 18:35:46", "opinion_text": "拟同意。", "detail": "", "implementation": "/"}},
          "countersign": [],
          "compliance": {{"department": "审计部/法务合规部", "person": "李莹", "time": "2026-05-09 10:51:09", "opinion_text": "已阅。拟同意。", "detail": "", "implementation": "/"}},
          "chief": {{"person": "胡鹏斌", "time": "2026-05-11 10:21:13", "opinion_text": "{merged_chief_text}", "detail": "{merged_chief_text}", "implementation": "/"}},
          "review_rows": [
            {{"review_time": "2026-05-11 10:21:13", "review_unit": "首席合规官", "review_opinion": "同意", "detail": "{merged_chief_text}", "implementation": "/"}}
          ],
          "warnings": []
        }}
        """
        source_text = """
签发意见
同意提交总办会审议。请履行会前传签程序。
中航建设直属 徐勤 2026-05-11 10:40:42
拟同意，建议提交总经理办公会会议审议。
中航建设直属 胡鹏斌 2026-05-11 10:21:13
会签
"""
        completions = _FakeCompletions([qwen_json, qwen_json])

        with patch("llm_client.get_llm_client", return_value=_FakeClient(completions)):
            item = extract_compliance_item(source_text, {"审计部/法务合规部": "李莹"})

        chief_rows = [row for row in item["review_rows"] if row["review_unit"] == "首席合规官"]

        self.assertEqual(len(chief_rows), 1)
        self.assertEqual(chief_rows[0]["review_time"], "2026-05-11 10:21:13")
        self.assertEqual(chief_rows[0]["review_opinion"], "同意")
        self.assertEqual(chief_rows[0]["detail"], "/")
        self.assertNotIn("请履行会前传签程序", json.dumps(chief_rows[0], ensure_ascii=False))

    def test_extract_compliance_item_writes_debug_snapshot(self):
        qwen_json = """
        {
          "title": "关于测试事项的请示",
          "procedure": "总办会审议",
          "attachments": [],
          "undertaking": {"department": "规划与资产部/深化改革领导小组办公室", "person": "富小鹏", "time": "2026-05-08 18:35:46", "opinion_text": "拟同意。", "detail": "", "implementation": "/"},
          "countersign": [],
          "compliance": {"department": "审计部/法务合规部", "person": "李莹", "time": "2026-05-09 10:51:09", "opinion_text": "已阅。拟同意。", "detail": "", "implementation": "/"},
          "chief": {"person": "胡鹏斌", "time": "2026-05-11 10:21:13", "opinion_text": "拟同意，建议提交总经理办公会议审议。", "detail": "", "implementation": "/"},
          "warnings": []
        }
        """
        completions = _FakeCompletions([qwen_json, qwen_json])
        TEST_TMP_ROOT.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmpdir:
            debug_path = Path(tmpdir) / "debug-last.json"
            with patch("llm_client.get_llm_client", return_value=_FakeClient(completions)), \
                 patch("utils.compliance_ledger.COMPLIANCE_DEBUG_PATH", debug_path):
                extract_compliance_item("拟同意，建议提交总经理办公会议审议。 中航建设直属 胡鹏斌 2026-05-11 10:21:13", {"审计部/法务合规部": "李莹"})

            debug_data = json.loads(debug_path.read_text(encoding="utf-8"))

        self.assertEqual(debug_data["debug_version"], "compliance-chief-diagnosis-v1")
        self.assertIn("chief_after_fix", debug_data)
        self.assertTrue(debug_data["final_review_rows"])


class ComplianceDetailForOpinionTests(unittest.TestCase):
    def test_short_agree_opinion_returns_slash(self):
        self.assertEqual(_detail_for_opinion({"opinion_text": "拟同意。"}), "/")

    def test_long_agree_opinion_preserves_content(self):
        long_opinion = "拟同意。另，会计师事务所在四川民航大厦2025年度财务决算审计报告的强调事项段，提醒财务报表使用者关注相关情况。"
        result = _detail_for_opinion({"opinion_text": long_opinion})
        self.assertEqual(result, long_opinion)

    def test_explicit_detail_field_takes_precedence(self):
        result = _detail_for_opinion({"detail": "明确的意见", "opinion_text": "拟同意。"})
        self.assertEqual(result, "明确的意见")

    def test_supplement_opinion_returns_full_text(self):
        text = "建议补充预算测算依据后再行审议"
        self.assertEqual(_detail_for_opinion({"opinion_text": text}), text)


class ComplianceIsComplianceDepartmentTests(unittest.TestCase):
    def test_configured_person_matches(self):
        persons = {"审计部/法务合规部": "李莹"}
        self.assertTrue(_is_compliance_department("审计部/法务合规部", "李莹", persons))

    def test_non_configured_person_in_same_dept_does_not_match(self):
        persons = {"审计部/法务合规部": "李莹"}
        self.assertFalse(_is_compliance_department("审计部/法务合规部", "宋媛媛", persons))
        self.assertFalse(_is_compliance_department("审计部/法务合规部", "申奇奇", persons))


class ComplianceSmokeTestPDFScenario(unittest.TestCase):
    """冒烟测试：模拟四川民航大厦股东会 PDF 的完整提取场景"""

    RESPONSIBLE_PERSONS = {
        "规划与资产部/深化改革领导小组办公室": "富小鹏",
        "财务部": "杨焕",
        "审计部/法务合规部": "李莹",
        "人力资源部": "陈锐",
        "党群办公室/董事会办公室/行政办公室": "刘芳",
        "安全质量部": "霍晓冬",
        "纪委办公室/巡察工作领导小组办公室": "边宁",
        "西南分公司（四川中航物业）": "张虎",
    }

    def test_full_scenario_produces_expected_review_rows(self):
        item = normalize_extracted_item({
            "title": "关于对四川民航大厦宾馆有限公司2026年股东会表决意见的请示",
            "procedure": "总办会审议",
            "attachments": ["会议通知", "四川民航宾馆股东会文件", "重大经营决策合规审查意见表"],
            "undertaking": {
                "department": "规划与资产部/深化改革领导小组办公室",
                "person": "富小鹏",
                "time": "2026-05-08 18:35:46",
                "opinion_text": "拟同意，请会签后呈胡总阅示。",
            },
            "approval_entries": [
                {"department": "审计部/法务合规部", "person": "宋媛媛", "time": "2026-05-09 19:01:35", "opinion_text": "拟同意。"},
                {"department": "财务部", "person": "杨焕", "time": "2026-05-09 15:08:46",
                 "opinion_text": "拟同意。另，会计师事务所在四川民航大厦2025年度财务决算审计报告的强调事项段，提醒财务报表使用者关注四川民航大厦2019年4月起停业等情况。"},
                {"department": "审计部/法务合规部", "person": "申奇奇", "time": "2026-05-09 14:44:59",
                 "opinion_text": "已阅核。已要求承办单位出具经本单位合规管理员签署的合规审查意见（后附）。综合规审查复核，本事项属于需经公司总经理办公会决策的重大经营事项，未发现存在重大合规风险。"},
                {"department": "审计部/法务合规部", "person": "李莹", "time": "2026-05-09 10:51:09",
                 "opinion_text": "请奇奇阅核、媛媛阅，明确是否为经理层决策事项，注意合规审查最新要求。"},
                {"department": "西南分公司（四川中航物业）直属", "person": "张虎", "time": "2026-05-09 09:53:27", "opinion_text": "已阅。拟同意。"},
                {"department": "西南分公司（四川中航物业）直属", "person": "王永君", "time": "2026-05-09 09:35:13", "opinion_text": "已阅"},
                {"department": "党群办公室/董事会办公室/行政办公室", "person": "刘芳", "time": "2026-05-09 09:49:20",
                 "opinion_text": "已核，按照重大事项权责清单（2026版），该事项属于经理层审议重大经营管理事项，建议经理层通过总办会审议方式行权。呈胡总、徐勤阅示。"},
                {"department": "中航建设直属", "person": "徐勤", "time": "2026-05-11 10:40:42",
                 "opinion_text": "同意提交总办会审议。请履行会前传签程序。"},
                {"department": "中航建设直属", "person": "胡鹏斌", "time": "2026-05-11 10:21:13",
                 "opinion_text": "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。"},
            ],
            "countersign": [],
            "compliance": {"department": "审计部/法务合规部", "person": "李莹", "time": "2026-05-09 10:51:09",
                           "opinion_text": "请奇奇阅核、媛媛阅，明确是否为经理层决策事项，注意合规审查最新要求。"},
            "chief": {"person": "胡鹏斌", "time": "2026-05-11 10:21:13",
                      "opinion_text": "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。"},
            "warnings": [],
        }, self.RESPONSIBLE_PERSONS)

        rows = item["review_rows"]
        units = [row["review_unit"] for row in rows]

        self.assertIn("首席合规官", units)
        self.assertIn("合规管理牵头部门（审计部/法务合规部）", units)
        self.assertIn("会签单位（财务部）", units)
        self.assertIn("会签单位（西南分公司（四川中航物业））", units)
        self.assertIn("会签单位（党群办公室/董事会办公室/行政办公室）", units)
        self.assertIn("承办单位（规划与资产部/深化改革领导小组办公室）", units)

        self.assertEqual(len(units), 6)

        # 非配置负责人不应出现
        all_text = json.dumps(rows, ensure_ascii=False)
        self.assertNotIn("宋媛媛", all_text)
        self.assertNotIn("申奇奇", all_text)
        self.assertNotIn("王永君", all_text)

        # 杨焕的长意见应保留在 detail 中
        finance_row = [r for r in rows if "财务部" in r["review_unit"]][0]
        self.assertIn("会计师事务所", finance_row["detail"])

        # 刘芳的长意见应保留在 detail 中
        party_row = [r for r in rows if "党群" in r["review_unit"]][0]
        self.assertIn("重大事项权责清单", party_row["detail"])

        # 首席合规官意见去重后 detail 应为 /（徐勤意见已被剥离）
        chief_row = [r for r in rows if r["review_unit"] == "首席合规官"][0]
        self.assertEqual(chief_row["review_opinion"], "同意")
        self.assertEqual(chief_row["detail"], "/")


class ComplianceFixChiefFromTextTests(unittest.TestCase):
    SIGN_SECTION = (
        "同意提交总办会审议。请履行会前传签程序。\n"
        "                                中航建设直属 徐勤 2026-05-11 10:40:42\n"
        "拟同意，建议提交总经理办公会议审议。\n"
        "                                中航建设直属 胡鹏斌 2026-05-11 10:21:13\n"
    )

    def test_corrects_merged_chief_opinion(self):
        raw = {
            "chief": {
                "person": "胡鹏斌",
                "time": "2026-05-11 10:21:13",
                "opinion_text": "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。",
            }
        }
        result = _fix_chief_opinion_from_text(raw, self.SIGN_SECTION)
        self.assertEqual(result["chief"]["opinion_text"], "拟同意，建议提交总经理办公会议审议。")

    def test_no_change_when_opinion_already_correct(self):
        raw = {
            "chief": {
                "person": "胡鹏斌",
                "opinion_text": "拟同意，建议提交总经理办公会议审议。",
            }
        }
        result = _fix_chief_opinion_from_text(raw, self.SIGN_SECTION)
        self.assertEqual(result["chief"]["opinion_text"], "拟同意，建议提交总经理办公会议审议。")

    def test_no_change_when_chief_not_found_in_text(self):
        raw = {
            "chief": {
                "person": "胡鹏斌",
                "opinion_text": "同意提交总办会审议。拟同意，建议提交。",
            }
        }
        result = _fix_chief_opinion_from_text(raw, "无关文本内容")
        self.assertEqual(result["chief"]["opinion_text"], "同意提交总办会审议。拟同意，建议提交。")

    def test_ignores_name_in_body_text_without_timestamp(self):
        """正文中提到胡鹏斌但没有时间戳的不应匹配，应找签发区带时间戳的那个"""
        text_with_body_mention = (
            "建议经理层通过总办会审议方式行权。呈胡鹏斌总、徐勤阅示。\n"
            "                党群办公室 刘芳 2026-05-11 09:49:20\n"
            "同意提交总办会审议。请履行会前传签程序。\n"
            "                中航建设直属 徐勤 2026-05-11 10:40:42\n"
            "拟同意，建议提交总经理办公会议审议。\n"
            "                中航建设直属 胡鹏斌 2026-05-11 10:21:13\n"
        )
        raw = {
            "chief": {
                "person": "胡鹏斌",
                "time": "2026-05-11 10:21:13",
                "opinion_text": "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。",
            }
        }
        result = _fix_chief_opinion_from_text(raw, text_with_body_mention)
        self.assertEqual(result["chief"]["opinion_text"], "拟同意，建议提交总经理办公会议审议。")

    def test_corrects_merged_chief_opinion_when_signature_area_has_no_line_breaks(self):
        raw = {
            "chief": {
                "person": "胡鹏斌",
                "time": "2026-05-11 10:21:13",
                "opinion_text": "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。",
                "detail": "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。",
            }
        }
        one_line_text = (
            "签发意见 同意提交总办会审议。请履行会前传签程序。"
            "中航建设直属 徐勤 2026-05-11 10:40:42 "
            "拟同意，建议提交总经理办公会议审议。中航建设直属 胡鹏斌 2026-05-11 10:21:13 会签"
        )
        result = _fix_chief_opinion_from_text(raw, one_line_text)
        self.assertEqual(result["chief"]["opinion_text"], "拟同意，建议提交总经理办公会议审议。")
        self.assertEqual(result["chief"]["detail"], "")

    def test_corrects_merged_chief_opinion_when_ocr_table_cell_is_already_merged(self):
        raw = {
            "chief": {
                "person": "胡鹏斌",
                "time": "2026-05-11 10:40:42",
                "opinion_text": "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。",
                "detail": "",
            }
        }
        merged_table_text = (
            "|签发|同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。|"
            "中航建设直属 徐勤 2026-05-11 10:40:42 中航建设直属 胡鹏斌 2026-05-11 10:40:42|"
        )
        result = _fix_chief_opinion_from_text(raw, merged_table_text)
        self.assertEqual(result["chief"]["opinion_text"], "拟同意，建议提交总经理办公会议审议。")
        self.assertEqual(result["chief"]["detail"], "")

    def test_corrects_merged_chief_opinion_when_chief_disagrees(self):
        raw = {
            "chief": {
                "person": "胡鹏斌",
                "time": "2026-05-11 10:40:42",
                "opinion_text": "同意提交总办会审议。请履行会前传签程序。不同意，建议暂缓提交。",
                "detail": "",
            }
        }
        merged_table_text = (
            "|签发|同意提交总办会审议。请履行会前传签程序。不同意，建议暂缓提交。|"
            "中航建设直属 徐勤 2026-05-11 10:40:42 中航建设直属 胡鹏斌 2026-05-11 10:40:42|"
        )
        result = _fix_chief_opinion_from_text(raw, merged_table_text)
        self.assertEqual(result["chief"]["opinion_text"], "不同意，建议暂缓提交。")
        self.assertEqual(result["chief"]["detail"], "")


class ComplianceDeduplicateChiefOpinionTests(unittest.TestCase):
    def test_strips_other_signer_opinion_from_chief(self):
        chief = {"person": "胡鹏斌", "opinion_text": "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。"}
        entries = [
            {"person": "徐勤", "opinion_text": "同意提交总办会审议。请履行会前传签程序。"},
            {"person": "胡鹏斌", "opinion_text": "同意提交总办会审议。请履行会前传签程序。拟同意，建议提交总经理办公会议审议。"},
        ]
        result = _deduplicate_chief_opinion(chief, entries)
        self.assertEqual(result["opinion_text"], "拟同意，建议提交总经理办公会议审议。")

    def test_no_change_when_no_overlap(self):
        chief = {"person": "胡鹏斌", "opinion_text": "拟同意，建议提交总经理办公会议审议。"}
        entries = [
            {"person": "徐勤", "opinion_text": "同意提交总办会审议。"},
            {"person": "胡鹏斌", "opinion_text": "拟同意，建议提交总经理办公会议审议。"},
        ]
        result = _deduplicate_chief_opinion(chief, entries)
        self.assertEqual(result["opinion_text"], "拟同意，建议提交总经理办公会议审议。")

    def test_no_change_when_opinion_too_short(self):
        chief = {"person": "胡鹏斌", "opinion_text": "同意"}
        entries = [{"person": "徐勤", "opinion_text": "已阅"}]
        result = _deduplicate_chief_opinion(chief, entries)
        self.assertEqual(result["opinion_text"], "同意")

    def test_strips_multiple_other_opinions(self):
        chief = {"person": "胡鹏斌", "opinion_text": "同意提交总办会审议。请履行会前传签程序。同意，请按规定程序办理。拟同意，建议提交总经理办公会议审议。"}
        entries = [
            {"person": "徐勤", "opinion_text": "同意提交总办会审议。请履行会前传签程序。"},
            {"person": "王总", "opinion_text": "同意，请按规定程序办理。"},
            {"person": "胡鹏斌", "opinion_text": "同意提交总办会审议。请履行会前传签程序。同意，请按规定程序办理。拟同意，建议提交总经理办公会议审议。"},
        ]
        result = _deduplicate_chief_opinion(chief, entries)
        self.assertEqual(result["opinion_text"], "拟同意，建议提交总经理办公会议审议。")


if __name__ == "__main__":
    unittest.main()
