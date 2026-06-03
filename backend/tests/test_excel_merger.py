# -*- coding: utf-8 -*-
"""三台账合并核心计算测试 (v3.6.18)。

excel_merger.py 是三表合并的核心，此前覆盖率仅 7%。本文件覆盖：
  * 纯函数：normalize_contract_no / clean_header / find_header_row_idx /
    build_index / determine_status（8 个状态分支）
  * IO 管线：read_file_as_records（表头探测 / 合同列定位 / 合计行跳过 /
    汇总 Sheet 优先 / 缺列报错）
  * 主管线：merge_ledgers（匹配统计 / 模糊匹配 / 一对多笛卡尔积 / 去重 /
    优先级排序 / 仅合同表 / 输出列与状态着色），通过回读输出 xlsx 验证。

纯内存 openpyxl，无网络 / 无磁盘 / 无 DB。
"""
from __future__ import annotations

import io
import sys
import unittest
from pathlib import Path

import openpyxl

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

from utils.excel_merger import (  # noqa: E402
    build_index,
    clean_header,
    determine_status,
    find_header_row_idx,
    merge_ledgers,
    normalize_contract_no,
    read_file_as_records,
)


def _xlsx(rows: list[list], *, sheet_name: str = "Sheet1",
          extra_sheets: list[tuple[str, list[list]]] | None = None) -> bytes:
    """把二维数组写成 xlsx bytes（第一行通常是表头）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in rows:
        ws.append(r)
    for name, srows in (extra_sheets or []):
        s = wb.create_sheet(name)
        for r in srows:
            s.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _read_rows(b: bytes) -> list[tuple]:
    wb = openpyxl.load_workbook(io.BytesIO(b), data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


# ── 纯函数 ────────────────────────────────────────────────────────────────────
class NormalizeContractNoTests(unittest.TestCase):
    def test_lowercases_and_strips(self):
        self.assertEqual(normalize_contract_no("  ABC-123  "), "abc-123")

    def test_fullwidth_parens_to_halfwidth(self):
        self.assertEqual(normalize_contract_no("（2022）ABC"), "(2022)abc")

    def test_trailing_hao_removed(self):
        self.assertEqual(normalize_contract_no("HT-2022-001号"), "ht-2022-001")

    def test_combined(self):
        self.assertEqual(normalize_contract_no("（2022）HT-001号"), "(2022)ht-001")

    def test_non_str_input(self):
        self.assertEqual(normalize_contract_no(12345), "12345")


class CleanHeaderTests(unittest.TestCase):
    def test_none_to_empty(self):
        self.assertEqual(clean_header(None), "")

    def test_strips_leading_asterisk_and_spaces(self):
        self.assertEqual(clean_header("  *合同编号 "), "合同编号")

    def test_number_header(self):
        self.assertEqual(clean_header(2022), "2022")


class FindHeaderRowIdxTests(unittest.TestCase):
    def test_header_on_first_row(self):
        b = _xlsx([["合同编号", "金额", "名称"], ["A1", 1, "x"]])
        wb = openpyxl.load_workbook(io.BytesIO(b))
        self.assertEqual(find_header_row_idx(wb.active), 1)

    def test_header_after_title_row(self):
        # 第1行只有标题（1 个非空），第2行是真正表头（3 个非空）
        b = _xlsx([["某某公司台账"], ["合同编号", "金额", "名称"], ["A1", 1, "x"]])
        wb = openpyxl.load_workbook(io.BytesIO(b))
        self.assertEqual(find_header_row_idx(wb.active), 2)


class BuildIndexTests(unittest.TestCase):
    def test_groups_by_normalized_key(self):
        records = [
            {"[采购] 合同编号": "（2022）A号"},
            {"[采购] 合同编号": "(2022)a"},
            {"[采购] 合同编号": "B-2"},
        ]
        idx = build_index(records, "[采购] 合同编号")
        self.assertEqual(len(idx["(2022)a"]), 2)
        self.assertEqual(len(idx["b-2"]), 1)

    def test_skips_empty_contract_no(self):
        records = [{"[采购] 合同编号": None}, {"[采购] 合同编号": "  "}, {"[采购] 合同编号": "X"}]
        idx = build_index(records, "[采购] 合同编号")
        self.assertEqual(list(idx.keys()), ["x"])


class DetermineStatusTests(unittest.TestCase):
    def test_only_contract_uploaded_is_full_match(self):
        self.assertEqual(determine_status(False, False, False, False), ("全部匹配", 1))

    def test_all_uploaded_and_matched(self):
        self.assertEqual(determine_status(True, True, True, True), ("全部匹配", 1))

    def test_purchase_only_unmatched_is_priority3(self):
        text, pri = determine_status(True, False, False, False)
        self.assertEqual(pri, 3)
        self.assertEqual(text, "缺少采购表内容")

    def test_partial_match_is_priority2(self):
        # 采购匹配上、财务没匹配上 → 部分匹配
        text, pri = determine_status(True, True, True, False)
        self.assertEqual(pri, 2)
        self.assertEqual(text, "缺少财务表内容")

    def test_both_unmatched_is_priority3(self):
        text, pri = determine_status(True, False, True, False)
        self.assertEqual(pri, 3)
        self.assertEqual(text, "缺少采购表和财务表内容")

    def test_finance_only_unmatched_priority3(self):
        text, pri = determine_status(False, False, True, False)
        self.assertEqual(pri, 3)
        self.assertEqual(text, "缺少财务表内容")


# ── read_file_as_records ──────────────────────────────────────────────────────
class ReadFileAsRecordsTests(unittest.TestCase):
    def test_basic_read_and_label_prefix(self):
        b = _xlsx([["合同编号", "金额"], ["HT-1", 100], ["HT-2", 200]])
        records, col = read_file_as_records(b, "合同")
        self.assertEqual(col, "[合同] 合同编号")
        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["[合同] 合同编号"], "HT-1")
        self.assertEqual(records[0]["[合同] 金额"], 100)

    def test_missing_contract_column_raises(self):
        b = _xlsx([["编号X", "金额"], ["1", 2]])
        with self.assertRaises(ValueError) as ctx:
            read_file_as_records(b, "合同")
        self.assertIn("合同编号", str(ctx.exception))

    def test_summary_and_empty_rows_skipped(self):
        b = _xlsx([
            ["合同编号", "金额"],
            ["HT-1", 100],
            [None, None],          # 全空行
            ["合计", 999],          # 合计行
            ["HT-2", 200],
        ])
        records, _ = read_file_as_records(b, "合同")
        nos = [r["[合同] 合同编号"] for r in records]
        self.assertEqual(nos, ["HT-1", "HT-2"])

    def test_prefers_summary_sheet(self):
        b = _xlsx(
            [["其它列"], ["x"]],
            sheet_name="明细",
            extra_sheets=[("汇总表", [["合同编号", "金额"], ["HT-9", 9]])],
        )
        records, col = read_file_as_records(b, "合同")
        self.assertEqual(col, "[合同] 合同编号")
        self.assertEqual(records[0]["[合同] 合同编号"], "HT-9")

    def test_header_with_asterisk_cleaned(self):
        b = _xlsx([["*合同编号", "*金额"], ["HT-1", 1]])
        records, col = read_file_as_records(b, "合同")
        self.assertEqual(col, "[合同] 合同编号")
        self.assertIn("[合同] 金额", records[0])


# ── merge_ledgers（主管线） ────────────────────────────────────────────────────
class MergeLedgersTests(unittest.TestCase):
    def test_contract_only_all_full_match(self):
        contract = _xlsx([["合同编号", "名称"], ["HT-1", "甲"], ["HT-2", "乙"]])
        out_bytes, stats = merge_ledgers(contract, None, None)
        self.assertEqual(stats["total_contract"], 2)
        self.assertEqual(stats["fully_matched"], 2)
        self.assertEqual(stats["unmatched"], 0)
        rows = _read_rows(out_bytes)
        self.assertEqual(rows[0][0], "匹配状态")
        self.assertEqual(rows[0][1], "合同编号")
        # 数据行状态均为"全部匹配"
        for r in rows[1:]:
            self.assertEqual(r[0], "全部匹配")

    def test_full_partial_unmatched_counts(self):
        contract = _xlsx([["合同编号", "名称"], ["HT-1", "甲"], ["HT-2", "乙"], ["HT-3", "丙"]])
        purchase = _xlsx([["合同编号", "采购额"], ["HT-1", 10], ["HT-2", 20]])
        finance = _xlsx([["合同编号", "付款"], ["HT-1", 100]])
        _, stats = merge_ledgers(contract, purchase, finance)
        self.assertEqual(stats["total_contract"], 3)
        self.assertEqual(stats["matched_purchase"], 2)  # HT-1, HT-2
        self.assertEqual(stats["matched_finance"], 1)   # HT-1
        self.assertEqual(stats["fully_matched"], 1)     # HT-1（采购+财务都匹配）
        self.assertEqual(stats["partial_matched"], 1)   # HT-2（仅采购匹配）
        self.assertEqual(stats["unmatched"], 1)         # HT-3（都没匹配）

    def test_fuzzy_matching_across_formats(self):
        # 合同表用全角括号+号，采购表用半角小写 → 应匹配
        contract = _xlsx([["合同编号", "名称"], ["（2022）HT-1号", "甲"]])
        purchase = _xlsx([["合同编号", "采购额"], ["(2022)ht-1", 10]])
        _, stats = merge_ledgers(contract, purchase, None)
        self.assertEqual(stats["matched_purchase"], 1)
        self.assertEqual(stats["fully_matched"], 1)

    def test_one_to_many_cartesian(self):
        # 一个合同对应两条采购记录 → 输出两行
        contract = _xlsx([["合同编号", "名称"], ["HT-1", "甲"]])
        purchase = _xlsx([["合同编号", "采购额"], ["HT-1", 10], ["HT-1", 20]])
        out_bytes, stats = merge_ledgers(contract, purchase, None)
        self.assertEqual(stats["matched_purchase"], 1)
        rows = _read_rows(out_bytes)
        self.assertEqual(len(rows) - 1, 2, "一对多应展开为 2 行")

    def test_identical_rows_deduped(self):
        # 两条完全相同的采购记录 → 去重后只剩 1 行
        contract = _xlsx([["合同编号", "名称"], ["HT-1", "甲"]])
        purchase = _xlsx([["合同编号", "采购额"], ["HT-1", 10], ["HT-1", 10]])
        out_bytes, _ = merge_ledgers(contract, purchase, None)
        rows = _read_rows(out_bytes)
        self.assertEqual(len(rows) - 1, 1, "完全一致的行应去重")

    def test_priority_sorting_full_before_unmatched(self):
        contract = _xlsx([["合同编号", "名称"], ["HT-miss", "无"], ["HT-1", "有"]])
        purchase = _xlsx([["合同编号", "采购额"], ["HT-1", 10]])
        out_bytes, _ = merge_ledgers(contract, purchase, None)
        rows = _read_rows(out_bytes)
        # 全部匹配（priority 1）排在缺失（priority 3）之前
        self.assertEqual(rows[1][0], "全部匹配")
        self.assertIn("缺少", rows[2][0])

    def test_output_columns_include_all_label_prefixes(self):
        contract = _xlsx([["合同编号", "名称"], ["HT-1", "甲"]])
        purchase = _xlsx([["合同编号", "采购额"], ["HT-1", 10]])
        finance = _xlsx([["合同编号", "付款"], ["HT-1", 100]])
        out_bytes, _ = merge_ledgers(contract, purchase, finance)
        headers = _read_rows(out_bytes)[0]
        self.assertIn("匹配状态", headers)
        self.assertIn("合同编号", headers)
        self.assertIn("[合同] 名称", headers)
        self.assertIn("[采购] 采购额", headers)
        self.assertIn("[财务] 付款", headers)


if __name__ == "__main__":
    unittest.main()
