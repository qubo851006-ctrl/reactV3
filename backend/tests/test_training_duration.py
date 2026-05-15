"""
培训时长（课时）功能单元测试 & 回归测试

覆盖范围：
  1. _calc_duration_hours  — 课时计算逻辑（含边界）
  2. _extract_training_time — LLM 时间提取（mock LLM，不发真实请求）
  3. excel_writer.append_record — 新建文件包含新列
  4. excel_writer._migrate_headers — 旧版文件自动迁移表头
"""
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))


# ─────────────────────────────────────────────────────────────
# 1. _calc_duration_hours
# ─────────────────────────────────────────────────────────────

class CalcDurationHoursTests(unittest.TestCase):
    def setUp(self):
        from routers.training import _calc_duration_hours
        self.calc = _calc_duration_hours

    def test_single_day_full_day(self):
        """09:00-17:00 单天 → 480 分钟 / 40 = 12.0 课时"""
        self.assertEqual(self.calc("09:00", "17:00", 1), 12.0)

    def test_single_day_half_session(self):
        """09:00-12:00 单天 → 180 分钟 / 40 = 4.5 课时"""
        self.assertEqual(self.calc("09:00", "12:00", 1), 4.5)

    def test_multi_day(self):
        """09:00-17:00 × 3 天 → 1440 分钟 / 40 = 36.0 课时"""
        self.assertEqual(self.calc("09:00", "17:00", 3), 36.0)

    def test_result_rounds_to_one_decimal(self):
        """非整除时保留 1 位小数：100 分钟 / 40 = 2.5"""
        self.assertEqual(self.calc("09:00", "10:40", 1), 2.5)

    def test_days_less_than_1_treated_as_1(self):
        """days=0 视作 1 天，不得返回 0"""
        result = self.calc("09:00", "17:00", 0)
        self.assertEqual(result, 12.0)

    def test_invalid_time_format_returns_zero(self):
        """时间格式非法时返回 0.0，不抛异常"""
        self.assertEqual(self.calc("9点", "下午5点", 1), 0.0)

    def test_empty_strings_return_zero(self):
        """空字符串返回 0.0"""
        self.assertEqual(self.calc("", "", 1), 0.0)

    def test_end_before_start_returns_zero(self):
        """结束时间早于开始时间（数据错误）返回 0.0"""
        self.assertEqual(self.calc("17:00", "09:00", 1), 0.0)


# ─────────────────────────────────────────────────────────────
# 2. _extract_training_time（mock LLM，不发网络请求）
# ─────────────────────────────────────────────────────────────

class ExtractTrainingTimeTests(unittest.TestCase):
    def _call_with_mock_llm(self, llm_reply: str, notice_text: str = "培训时间：09:00-17:00"):
        """patch OpenAI 客户端，让 LLM 返回指定文本。
        注意：_extract_training_time 在函数体内 import openai，须 patch openai.OpenAI。
        """
        mock_response = MagicMock()
        mock_response.choices[0].message.content = llm_reply

        mock_client_instance = MagicMock()
        mock_client_instance.chat.completions.create.return_value = mock_response

        with patch("openai.OpenAI", return_value=mock_client_instance), \
             patch("httpx.Client", return_value=MagicMock()), \
             patch.dict("os.environ", {
                 "AIRCHINA_API_KEY": "fake-key",
                 "AIRCHINA_BASE_URL": "http://fake",
                 "MODEL_CLASSIFY": "test-model",
                 "AI_HTTP_VERIFY_SSL": "false",
                 "AI_HTTP_HOST": "",
             }):
            from routers.training import _extract_training_time
            return _extract_training_time(notice_text)

    def test_extracts_single_day(self):
        """LLM 返回标准格式 → 正确解析单天结果"""
        result = self._call_with_mock_llm("开始时间：09:00\n结束时间：17:00\n天数：1")
        self.assertEqual(result["start_time"], "09:00")
        self.assertEqual(result["end_time"], "17:00")
        self.assertEqual(result["days"], 1)
        self.assertEqual(result["duration_hours"], 12.0)

    def test_extracts_multi_day(self):
        """多天培训 → 课时按天数累乘"""
        result = self._call_with_mock_llm("开始时间：09:00\n结束时间：17:00\n天数：3")
        self.assertEqual(result["days"], 3)
        self.assertEqual(result["duration_hours"], 36.0)

    def test_zero_pads_single_digit_hour(self):
        """9:00 → 补零为 09:00"""
        result = self._call_with_mock_llm("开始时间：9:00\n结束时间：17:00\n天数：1")
        self.assertEqual(result["start_time"], "09:00")

    def test_empty_notice_returns_empty(self):
        """空通知文字直接返回空结果，不调用 LLM"""
        from routers.training import _extract_training_time
        result = _extract_training_time("   ")
        self.assertEqual(result["start_time"], "")
        self.assertEqual(result["end_time"], "")
        self.assertEqual(result["duration_hours"], 0.0)

    def test_llm_missing_time_returns_empty(self):
        """LLM 回复中无时间信息 → 返回空字符串和 0.0"""
        result = self._call_with_mock_llm("开始时间：\n结束时间：\n天数：1")
        self.assertEqual(result["start_time"], "")
        self.assertEqual(result["end_time"], "")
        self.assertEqual(result["duration_hours"], 0.0)

    def test_llm_exception_returns_empty(self):
        """LLM 调用异常时返回空结果，不向外抛异常"""
        with patch("openai.OpenAI", side_effect=Exception("network error")), \
             patch("httpx.Client", return_value=MagicMock()), \
             patch.dict("os.environ", {"AIRCHINA_API_KEY": "k", "AIRCHINA_BASE_URL": "u"}):
            from routers.training import _extract_training_time
            result = _extract_training_time("09:00-17:00 培训")
        self.assertEqual(result["duration_hours"], 0.0)


# ─────────────────────────────────────────────────────────────
# 3. excel_writer — 新建文件包含新列
# ─────────────────────────────────────────────────────────────

class ExcelWriterNewFileTests(unittest.TestCase):
    def test_new_file_has_duration_column(self):
        """新创建的 Excel 文件表头必须包含「培训时长（课时）」"""
        import openpyxl
        from utils.excel_writer import HEADERS, append_record

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "培训统计表.xlsx"
            with patch("utils.excel_writer.EXCEL_PATH", str(path)):
                append_record(
                    date="2026-05-11",
                    topic="测试培训",
                    location="会议室",
                    department="法律部",
                    count=10,
                    category="合规培训",
                    archive_path="/data/archive/test",
                    duration_hours=12.0,
                )
            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            self.assertIn("培训时长（课时）", headers)

    def test_duration_hours_written_to_correct_column(self):
        """写入的课时值应出现在「培训时长（课时）」列对应位置"""
        import openpyxl
        from utils.excel_writer import append_record

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "培训统计表.xlsx"
            with patch("utils.excel_writer.EXCEL_PATH", str(path)):
                append_record(
                    date="2026-05-11",
                    topic="课时测试",
                    location="会议室",
                    department="",
                    count=5,
                    category="合规培训",
                    archive_path="/data/test",
                    duration_hours=4.5,
                )
            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            col_idx = headers.index("培训时长（课时）") + 1
            written_value = ws.cell(row=2, column=col_idx).value
            self.assertEqual(written_value, 4.5)

    def test_zero_duration_written_as_empty(self):
        """duration_hours=0.0 时单元格为空（openpyxl 读取空字符串返回 None）"""
        import openpyxl
        from utils.excel_writer import append_record

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "培训统计表.xlsx"
            with patch("utils.excel_writer.EXCEL_PATH", str(path)):
                append_record(
                    date="2026-05-11",
                    topic="无时长",
                    location="",
                    department="",
                    count=3,
                    category="合规培训",
                    archive_path="/data/test",
                    duration_hours=0.0,
                )
            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            col_idx = headers.index("培训时长（课时）") + 1
            cell_value = ws.cell(row=2, column=col_idx).value
            # openpyxl 将写入的空字符串读回为 None，两种情况均视为"空"
            self.assertFalse(cell_value, f"期望空单元格，实际值为 {cell_value!r}")

    def test_duplicate_training_updates_existing_row_by_date_and_topic(self):
        """同一培训日期+主题再次写入时更新原行，不新增重复记录"""
        import openpyxl
        from utils.excel_writer import append_record

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "培训统计表.xlsx"
            with patch("utils.excel_writer.EXCEL_PATH", str(path)):
                append_record(
                    date="2026-05-11",
                    topic="合规专题培训",
                    location="第一会议室",
                    department="法律部",
                    count=10,
                    category="合规培训",
                    archive_path="/archive/old",
                    duration_hours=4.0,
                )
                append_record(
                    date="2026-05-11",
                    topic=" 合规专题培训 ",
                    location="第二会议室",
                    department="审计部",
                    count=12,
                    category="法律培训",
                    archive_path="/archive/new",
                    duration_hours=8.0,
                )

            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            self.assertEqual(ws.max_row, 2)
            self.assertEqual(ws.cell(row=2, column=1).value, 1)
            self.assertEqual(ws.cell(row=2, column=4).value, "第二会议室")
            self.assertEqual(ws.cell(row=2, column=5).value, "审计部")
            self.assertEqual(ws.cell(row=2, column=6).value, 12)
            self.assertEqual(ws.cell(row=2, column=7).value, 8.0)
            self.assertEqual(ws.cell(row=2, column=9).value, "/archive/new")


# ─────────────────────────────────────────────────────────────
# 4. _migrate_headers — 旧版文件自动迁移（回归测试）
# ─────────────────────────────────────────────────────────────

class MigrateHeadersTests(unittest.TestCase):
    def _make_old_excel(self, path: Path):
        """创建缺少「培训时长（课时）」列的旧版 Excel 文件"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        old_headers = ["序号", "培训日期", "培训主题", "培训地点", "主办部门",
                        "参与人数", "培训类别", "归档路径", "录入时间"]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "培训记录"
        for col, h in enumerate(old_headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")
        # 写一条旧数据
        ws.append([1, "2026-01-01", "旧培训", "会议室", "法律部", 20, "合规培训", "/archive/old", "2026-01-01 09:00"])
        wb.save(str(path))

    def test_migrate_inserts_duration_column(self):
        """旧版文件写入新记录后，表头应出现「培训时长（课时）」"""
        import openpyxl
        from utils.excel_writer import append_record

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "培训统计表.xlsx"
            self._make_old_excel(path)

            with patch("utils.excel_writer.EXCEL_PATH", str(path)):
                append_record(
                    date="2026-05-11",
                    topic="新培训",
                    location="会议室",
                    department="法律部",
                    count=15,
                    category="合规培训",
                    archive_path="/data/new",
                    duration_hours=8.0,
                )

            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            self.assertIn("培训时长（课时）", headers)

    def test_migrate_preserves_old_data(self):
        """迁移后旧数据行不丢失，参与人数仍正确"""
        import openpyxl
        from utils.excel_writer import append_record

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "培训统计表.xlsx"
            self._make_old_excel(path)

            with patch("utils.excel_writer.EXCEL_PATH", str(path)):
                append_record(
                    date="2026-05-11",
                    topic="新培训",
                    location="",
                    department="",
                    count=5,
                    category="合规培训",
                    archive_path="/data/new",
                    duration_hours=4.0,
                )

            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            count_col = headers.index("参与人数") + 1
            # 旧数据行（row=2）的参与人数应仍为 20
            self.assertEqual(ws.cell(row=2, column=count_col).value, 20)

    def test_migrate_does_not_duplicate_column(self):
        """已迁移过的文件再次写入，不会重复插入「培训时长（课时）」列"""
        import openpyxl
        from utils.excel_writer import append_record

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "培训统计表.xlsx"

            with patch("utils.excel_writer.EXCEL_PATH", str(path)):
                # 第一次写入（新建文件，已有新列）
                append_record("2026-05-11", "培训A", "", "", 5, "合规", "/a", 4.0)
                # 第二次写入（不应重复插列）
                append_record("2026-05-11", "培训B", "", "", 8, "法律", "/b", 8.0)

            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            count = headers.count("培训时长（课时）")
            self.assertEqual(count, 1)


if __name__ == "__main__":
    unittest.main()
